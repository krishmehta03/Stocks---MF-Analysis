import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

def upgrade_excel():
    input_file = "Stocks & MF Analysis.xlsx"
    output_file = "Stocks & MF Analysis_V3.xlsx"
    
    print(f"Loading {input_file}...")
    wb = openpyxl.load_workbook(input_file)
    
    # Header styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 1. UPGRADE STOCK SHEET
    if "📥 Stock Data Input" in wb.sheetnames:
        ws_stock = wb["📥 Stock Data Input"]
        print("Upgrading Stock Data Input sheet...")
        
        # New column headers at row 3
        # Existing columns go up to O (column 15)
        new_headers_stock = [
            (16, 'P3', 'Dividends\nReceived (₹)'),
            (17, 'Q3', 'Tax Flag'),
            (18, 'R3', 'Total Return\n(w/ Div)'),
        ]
        
        for col_idx, cell_ref, value in new_headers_stock:
            cell = ws_stock[cell_ref]
            cell.value = value
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            
        # Add formulas for the new columns (Rows 4 to 50 for a start)
        for row in range(4, 51):
            # Tax Flag: Based on Buy Date in F
            ws_stock[f'Q{row}'] = f'=IF(ISBLANK(F{row}), "", IF(TODAY()-F{row}>365, "LTCG", "STCG"))'
            
            # Total Return: P&L (L) + Dividends (P)
            ws_stock[f'R{row}'] = f'=IF(ISBLANK(L{row}), "", L{row}+IF(ISNUMBER(P{row}), P{row}, 0))'

    # 2. UPGRADE MF SHEET
    if "📥 MF Data Input" in wb.sheetnames:
        ws_mf = wb["📥 MF Data Input"]
        print("Upgrading MF Data Input sheet...")
        
        # New column headers at row 3
        # Existing columns go up to P (column 16)
        new_headers_mf = [
            (17, 'Q3', 'Tax Flag'),
        ]
        
        for col_idx, cell_ref, value in new_headers_mf:
            cell = ws_mf[cell_ref]
            cell.value = value
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            
        # Add formulas for the new columns (Rows 4 to 50)
        for row in range(4, 51):
            # Tax Flag: Based on Buy Date in G. Equity MF LTCG is 1 yr, Debt is 3 yrs. We'll simplify to 1 yr for now.
            ws_mf[f'Q{row}'] = f'=IF(ISBLANK(G{row}), "", IF(TODAY()-G{row}>365, "LTCG", "STCG"))'

    print(f"Saving to {output_file}...")
    wb.save(output_file)
    print("Done!")

if __name__ == "__main__":
    upgrade_excel()
