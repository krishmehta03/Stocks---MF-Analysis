from flask import Flask, jsonify, request, render_template
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import requests
import time
import os
import json
import threading
import csv
import io
from datetime import datetime
import re

import shutil
import yfinance as yf
# Monkeypatch yfinance to use a modern browser User-Agent instead of ancient Chrome 39
yf.data.YfData.user_agent_headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36'
}
import pandas as pd
import hashlib
from dotenv import load_dotenv
from lib.supabase import get_supabase, get_supabase_admin

load_dotenv()

import razorpay

razorpay_client = None

def get_razorpay_client():
    global razorpay_client
    key_id = os.environ.get('RAZORPAY_KEY_ID')
    key_secret = os.environ.get('RAZORPAY_KEY_SECRET')
    if key_id:
        key_id = key_id.strip().strip("'").strip('"')
    if key_secret:
        key_secret = key_secret.strip().strip("'").strip('"')
    
    if key_id and key_secret:
        if not razorpay_client:
            try:
                razorpay_client = razorpay.Client(auth=(key_id, key_secret))
            except Exception as e:
                print(f"[Razorpay] Initialization error: {e}")
        return razorpay_client, key_id, key_secret
    return None, None, None

# Try to initialize at startup
try:
    get_razorpay_client()
except Exception as e:
    print(f"[Razorpay] Startup init error: {e}")


def get_current_user_id():
    auth_header = request.headers.get("Authorization")
    if not auth_header or not auth_header.startswith("Bearer "):
        return None
    token = auth_header.split(" ")[1]
    try:
        supabase = get_supabase_admin()
        response = supabase.auth.get_user(token)
        if response and response.user:
            return response.user.id
    except Exception as e:
        print(f"Auth error: {e}")
    return None

def hash_pin(pin):
    if pin is None:
        return None
    pin_str = str(pin).strip()
    if not pin_str:
        return None
    return hashlib.sha256(pin_str.encode('utf-8')).hexdigest()

PLAN_LIMITS = {
    "free": {
        "max_stocks": 10,
        "max_mf": 3,
        "live_price_update": False,
        "alerts": False,
        "ai_advisor": False,
        "sector_contribution": False,
        "benchmark_comparison": False,
        "tax_reports": False,
        "export_csv": False,
        "multiple_profiles": False
    },
    "pro": {
        "max_stocks": -1,  # unlimited
        "max_mf": -1,      # unlimited
        "live_price_update": True,
        "alerts": True,
        "ai_advisor": True,
        "sector_contribution": True,
        "benchmark_comparison": True,
        "tax_reports": True,
        "export_csv": True,
        "multiple_profiles": True
    },
    "wealth": {
        "max_stocks": -1,
        "max_mf": -1,
        "live_price_update": True,
        "alerts": True,
        "ai_advisor": True,
        "sector_contribution": True,
        "benchmark_comparison": True,
        "tax_reports": True,
        "export_csv": True,
        "multiple_profiles": True
    }
}

def get_user_plan(user_id: str) -> str:
    """Get user's current plan from Supabase, checking for expiry and auto downgrading if expired (uses admin client to bypass RLS)"""
    try:
        supabase = get_supabase_admin()
        result = supabase.table('profiles')\
            .select('plan, plan_expires_at')\
            .eq('id', user_id)\
            .single()\
            .execute()
        
        if not result.data:
            return 'free'
            
        plan = result.data.get('plan', 'free')
        expires_at = result.data.get('plan_expires_at')
        
        # Check expiry
        if expires_at and plan == 'pro':
            from datetime import timezone
            expiry = datetime.fromisoformat(
                expires_at.replace('Z', '+00:00'))
            if datetime.now(timezone.utc) > expiry:
                # Auto downgrade to free
                supabase.table('profiles')\
                    .update({'plan': 'free'})\
                    .eq('id', user_id)\
                    .execute()
                print(f"[get_user_plan] Plan for user {user_id} expired at {expires_at}. Downgraded to free.")
                return 'free'
        
        return plan
    except Exception as e:
        print(f"[get_user_plan] ERROR for user_id={user_id}: {e}")
        return 'free'

def check_plan_limit(
    user_id: str, 
    feature: str
) -> bool:
    """
    Returns True if user can access feature.
    Returns False if feature is locked.
    """
    plan = get_user_plan(user_id)
    limits = PLAN_LIMITS.get(plan, 
                             PLAN_LIMITS['free'])
    return limits.get(feature, False)

def check_stock_limit(user_id: str) -> dict:
    """
    Check if user can add more stocks.
    Returns {allowed: bool, current: int, 
             max: int, plan: str}
    """
    plan = get_user_plan(user_id)
    limits = PLAN_LIMITS.get(plan, 
                             PLAN_LIMITS['free'])
    max_stocks = limits['max_stocks']
    
    if max_stocks == -1:
        return {
            "allowed": True,
            "current": 0,
            "max": -1,
            "plan": plan
        }
    
    from lib.supabase_data import get_user_stock_holdings
    current = len(get_user_stock_holdings(user_id))
    
    return {
        "allowed": current < max_stocks,
        "current": current,
        "max": max_stocks,
        "plan": plan
    }


app = Flask(__name__)

# Multi-Profile System File Resolver
def get_profile_files():
    try:
        meta_file = "profiles_config.json"
        if not os.path.exists(meta_file):
            initial_meta = {
                "active_profile": "Default Portfolio",
                "profiles": {
                    "Default Portfolio": {
                        "excel_path": "Stocks & MF Analysis_V3.xlsx",
                        "config_path": "config.json"
                    }
                }
            }
            with open(meta_file, "w") as f:
                json.dump(initial_meta, f, indent=4)
        
        with open(meta_file, "r") as f:
            meta = json.load(f)
            
        active = meta.get("active_profile", "Default Portfolio")
        profile_info = meta.get("profiles", {}).get(active, {
            "excel_path": "Stocks & MF Analysis_V3.xlsx",
            "config_path": "config.json"
        })
        excel_path = profile_info.get("excel_path")
        config_path = profile_info.get("config_path")
        if excel_path:
            excel_path = os.path.normpath(excel_path.replace('\\', '/'))
        if config_path:
            config_path = os.path.normpath(config_path.replace('\\', '/'))
        return excel_path, config_path, active
    except Exception as e:
        print("Error resolving profile path:", e)
        return "Stocks & MF Analysis_V3.xlsx", "config.json", "Default Portfolio"

def load_profile_globals():
    global EXCEL_FILE, CONFIG_FILE, ACTIVE_PROFILE
    excel, config, active = get_profile_files()
    EXCEL_FILE = excel
    CONFIG_FILE = config
    ACTIVE_PROFILE = active
    return excel, config, active

# Global profile trackers
EXCEL_FILE = "Stocks & MF Analysis_V3.xlsx"
CONFIG_FILE = "config.json"
ACTIVE_PROFILE = "Default Portfolio"
load_profile_globals()


# Global state for background live price update progress
price_update_state = {
    "status": "idle",  # idle, running, completed, error, background_running, locked
    "current_index": 0,
    "total_tickers": 0,
    "current_ticker": "",
    "logs": [],
    "start_time": None,
    "last_failed_time": None,
    "last_updated_by_profile": {}  # active profile -> timestamp
}

def get_user_config_path(user_id):
    if not user_id:
        return CONFIG_FILE
    os.makedirs("profiles", exist_ok=True)
    safe_uid = "".join([c if c.isalnum() else "_" for c in str(user_id)])
    return f"profiles/config_{safe_uid}.json"

def load_config(user_id=None):
    path = get_user_config_path(user_id)
    if not os.path.exists(path):
        if os.path.exists(CONFIG_FILE) and path != CONFIG_FILE:
            try:
                shutil.copy(CONFIG_FILE, path)
            except Exception:
                pass
    
    if not os.path.exists(path):
        return {}
    try:
        with open(path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        # Ensure risk_alerts is enabled by default as requested by user
        if "widgets" not in data:
            data["widgets"] = {}
        if data["widgets"].get("risk_alerts") is False or "risk_alerts" not in data["widgets"]:
            data["widgets"]["risk_alerts"] = True
            with open(path, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=2)
        return data
    except Exception as e:
        print(f"Error loading/updating config file at {path}: {e}")
        try:
            with open(path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            return {}

def save_config(config, user_id=None):
    path = get_user_config_path(user_id)
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(config, f, indent=2)

def parse_date(date_val):
    if not date_val:
        return None
    if isinstance(date_val, datetime):
        return date_val
    for fmt in ('%d-%m-%Y', '%Y-%m-%d', '%d/%m/%Y', '%Y/%m/%d', '%d %b %Y', '%d %B %Y'):
        try:
            return datetime.strptime(str(date_val).strip(), fmt)
        except ValueError:
            pass
    return None

def format_date_str(date_val):
    d = parse_date(date_val)
    return d.strftime('%d %b %Y') if d else ""

def get_tax_flag(buy_date, threshold_days):
    d = parse_date(buy_date)
    if not d:
        return ""
    delta = datetime.now() - d
    return "LTCG" if delta.days > threshold_days else "STCG"

def get_holding_period_years(buy_date_raw) -> float:
    """Return holding period as decimal years (e.g. 1.3, 0.16). Returns 0.0 if no date."""
    d = parse_date(buy_date_raw)
    if not d:
        return 0.0
    return round((datetime.now() - d).days / 365.25, 2)

def compute_xirr(invested: float, current_value: float, buy_date_raw) -> float | None:
    """
    Compute XIRR for a simple 2-cashflow case:
      - Outflow: -invested on buy_date
      - Inflow:  +current_value today
    Returns annualised % or None if holding period < 7 days or invested == 0.
    Uses Newton-Raphson iteration — no external dependencies needed.
    """
    if not invested or invested <= 0:
        return None
    d = parse_date(buy_date_raw)
    if not d:
        return None
    days = (datetime.now() - d).days
    if days < 7:
        return None
    years = days / 365.25
    # Simple CAGR formula is a valid XIRR approximation for single-tranche investments
    try:
        xirr = (current_value / invested) ** (1.0 / years) - 1.0
        return round(xirr * 100, 2)
    except Exception:
        return None

def compute_mf_tax(pnl: float, tax_flag: str) -> float:
    """
    Estimate redemption tax liability for Indian MF taxation rules:
      - STCG: 20% on gains (post-2024 budget)
      - LTCG: 12.5% on gains exceeding ₹1,25,000 annual exemption
    Note: This is a per-holding estimate; actual tax depends on total annual gains.
    """
    if pnl <= 0:
        return 0.0
    if tax_flag == "STCG":
        return round(pnl * 0.20, 2)
    elif tax_flag == "LTCG":
        taxable = max(0, pnl - 125000)
        return round(taxable * 0.125, 2)
    return 0.0

# ─── In-memory caches (per server session) ────────────────────────────────────
_ticker_cache = {}   # "COMPANY NAME|NSE" -> "TATAMOTORS.NS"  (resolved yf symbol)
_info_cache   = {}   # "TATAMOTORS.NS"    -> {price, sector, ...}
_drawdown_cache = {} # "TATAMOTORS.NS|BUY_PRICE" -> {"days": int, "ts": float}
_mf_scheme_cache = {} # "FUND NAME" -> scheme_code (int)

_CACHE_FILE = "yfinance_cache.json"

def load_cache():
    global _ticker_cache, _info_cache, _drawdown_cache, _mf_scheme_cache
    if os.path.exists(_CACHE_FILE):
        try:
            with open(_CACHE_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
                _ticker_cache.update(data.get("ticker_cache", {}))
                _info_cache.update(data.get("info_cache", {}))
                _drawdown_cache.update(data.get("drawdown_cache", {}))
                _mf_scheme_cache.update(data.get("mf_scheme_cache", {}))
                # Filter out null values so we retry resolving them
                _ticker_cache = {k: v for k, v in _ticker_cache.items() if v is not None}
        except Exception as e:
            print("Error loading yfinance cache:", e)

def save_cache():
    try:
        with open(_CACHE_FILE, "w", encoding="utf-8") as f:
            json.dump({
                "ticker_cache": _ticker_cache,
                "info_cache": _info_cache,
                "drawdown_cache": _drawdown_cache,
                "mf_scheme_cache": _mf_scheme_cache
            }, f, indent=2)
    except Exception as e:
        print("Error saving yfinance cache:", e)

load_cache()

# Yahoo Finance sector label → simplified display name
_SECTOR_MAP = {
    "Technology":             "Technology",
    "Financial Services":     "Banking & Finance",
    "Consumer Cyclical":      "Consumer Discretionary",
    "Consumer Defensive":     "Consumer Staples",
    "Healthcare":             "Healthcare",
    "Basic Materials":        "Materials",
    "Industrials":            "Industrials",
    "Energy":                 "Energy",
    "Utilities":              "Utilities",
    "Real Estate":            "Real Estate",
    "Communication Services": "Communication",
}

_BROAD_SECTOR_MAP = {
    "automobiles - passenger cars": "Auto",
    "auto - trucks": "Auto",
    "steel/sponge iron/pig iron": "Metals",
    "metal - non ferrous": "Metals",
    "mining & minerals": "Metals",
    "power generation/distribution": "Power",
    "diamond & gold jewellery": "Consumer",
    "diamond  & gold  jewellery": "Consumer", # Double spaces
    "etfs": "ETF",
    "materials": "Materials",
    "utilities": "Utilities",
    "textile": "Textile",
    "bank - public": "Financials",
    "bank - private": "Financials",
    "finance/nbfc": "Financials",
    "banking & finance": "Financials",
    "financial services": "Financials",
    "financials": "Financials",
    "telecommunication - service provider": "Communication",
    "telecommunication - service  provider": "Communication",
    "electric equipment": "Industrials",
    "industrials": "Industrials",
    "energy": "Energy",
    "real estate": "Real Estate",
    "communication services": "Communication",
    "communication": "Communication",
    "technology": "Technology",
    "healthcare": "Healthcare",
    "other": "ETF"
}

def get_broad_sector(granular_sector: str) -> str:
    if not granular_sector:
        return "ETF"
    norm = granular_sector.strip().lower()
    return _BROAD_SECTOR_MAP.get(norm, granular_sector.strip())


def resolve_stock_sector(raw_sector: str, sym: str | None) -> tuple:
    """
    Resolve the (sector, industry, is_sector_missing) tuple for a stock.
    If the raw sector in Excel is missing, 'Other', or 'ETFs', try to look it up
    from yfinance cache or yfinance info.
    """
    raw_clean = str(raw_sector or "").strip()
    if raw_clean and raw_clean not in ("Other", "ETFs"):
        return get_broad_sector(raw_clean), raw_clean, False
        
    # Try resolving from yfinance
    if sym:
        cached_info = _info_cache.get(sym)
        fetched_sector = None
        if cached_info and cached_info.get("sector"):
            fetched_sector = cached_info["sector"]
        else:
            try:
                info = _yf_info(sym)
                fetched_sector = info.get("sector")
            except Exception:
                pass
                
        if fetched_sector:
            industry = _SECTOR_MAP.get(fetched_sector, fetched_sector)
            return get_broad_sector(industry), industry, False
            
    # Default to ETFs/ETF
    return "ETF", "ETFs", True




_CUSTOM_TICKER_MAP = {
    "NIPPON INDIA LTG": "LTGILTBEES.NS",
    "NIPPON INDIA NM": "MID150BEES.NS",
    "NIP. INDIA NIFTY": "NIFTYBEES.NS",
    "NIPPON INDIA NIFTY": "NIFTYBEES.NS",
    "NIP. INDIA GOLD": "GOLDBEES.NS",
    "NIPPON INDIA GOLD": "GOLDBEES.NS",
    "NASDAQ 100": "MON100.NS",
    "HDFC NSC 250 ETF": "HDFCSML250.NS",
    "ST BK OF INDIA": "SBIN.NS",
    "STATE BANK OF INDIA": "SBIN.NS",
    "POWER GRID CORPN": "POWERGRID.NS",
    "POWER GRID": "POWERGRID.NS",
    "POWER GRID CORPORATION": "POWERGRID.NS",
    "A R C FINANCE": "ARCFIN.BO",
    "ARC FINANCE": "ARCFIN.BO",
    "SERVOTECH POWER": "SERVOTECH.NS",
    "SERVOTECH POWER SYSTEMS": "SERVOTECH.NS",
    "VISAGAR POLYTEX": "VIVIDHA.NS",
    "CPSE ETF*": "CPSEETF.NS",
    "CPSE ETF": "CPSEETF.NS",
    "TATA MOTORS": "TMCV.NS",
    "TATA MOTORS LIMITED": "TMCV.NS",
    "TATA MOTORS PASS VEH LTD": "TMPV.NS",
    "TATA MOTORS PASSENGER VEHICLES": "TMPV.NS",
    "TATA STEEL": "TATASTEEL.NS",
    "TATA STEEL LIMITED": "TATASTEEL.NS",
    "ASHOK LEYLAND": "ASHOKLEY.NS",
    "ASHOK LEYLAND LIMITED": "ASHOKLEY.NS",
    "COAL INDIA": "COALINDIA.NS",
    "COAL INDIA LTD": "COALINDIA.NS",
    "COAL INDIA LIMITED": "COALINDIA.NS",
    "TITAN COMPANY": "TITAN.NS",
    "TITAN COMPANY LIMITED": "TITAN.NS",
    "VEDANTA": "VEDL.NS",
    "VEDANTA LIMITED": "VEDL.NS",
    "VODAFONE IDEA": "IDEA.NS",
    "VODAFONE IDEA LIMITED": "IDEA.NS",
    "NHPC LTD": "NHPC.NS",
    "NHPC LIMITED": "NHPC.NS",
    "VAML": "VAML.NS",
    "S A I L": "SAIL.NS",
    "HDFC BANK": "HDFCBANK.NS",
}


def _clean_search_query(q: str) -> str:
    q = q.strip()
    q = re.sub(r'(?i)\bst\s+bk\b', 'State Bank', q)
    q = re.sub(r'(?i)\bnip\.?\b', 'Nippon', q)
    q = re.sub(r'(?i)\b(ltd|limited|corp|corpn|corporation|co|company)\b', '', q)
    q = re.sub(r'\s+', ' ', q).strip()
    return q


def _resolve_ticker(company_name: str, exchange: str) -> str | None:
    """
    Convert a full company name (e.g. 'TATA MOTORS LIMITED') to the correct
    Yahoo Finance ticker symbol (e.g. 'TATAMOTORS.NS') using direct search requests.
    Results are cached so each name is resolved only once per session.
    """
    name_clean = company_name.strip()
    clean_upper = name_clean.upper()

    # Strategy 0: Direct custom lookup mapping
    if clean_upper in _CUSTOM_TICKER_MAP:
        resolved = _CUSTOM_TICKER_MAP[clean_upper]
        cache_key = f"{clean_upper}|{exchange.strip().upper()}"
        if _ticker_cache.get(cache_key) != resolved:
            _ticker_cache[cache_key] = resolved
            save_cache()
        return resolved

    cache_key = f"{clean_upper}|{exchange.strip().upper()}"
    if cache_key in _ticker_cache:
        return _ticker_cache[cache_key]

    ex_suffix = ".BO" if "BSE" in exchange.upper() else ".NS"

    # Strategy 1 — treat the stored value as a raw ticker (works if user typed TATAMOTORS)
    direct = f"{name_clean}{ex_suffix}"
    try:
        info = yf.Ticker(direct).fast_info
        # fast_info raises on invalid tickers; check price is non-zero
        if getattr(info, 'last_price', None):
            _ticker_cache[cache_key] = direct
            save_cache()
            return direct
    except Exception:
        pass

    # Strategy 2 — use Yahoo Finance Search API directly
    try:
        cleaned_query = _clean_search_query(name_clean)
        url = "https://query2.finance.yahoo.com/v1/finance/search"
        params = {"q": cleaned_query, "quotesCount": 5, "newsCount": 0}
        headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}
        r = requests.get(url, params=params, headers=headers, timeout=5)
        if r.status_code == 200:
            results = r.json().get("quotes", [])
            # Prefer exact exchange match, then any Indian exchange
            for preference in (ex_suffix.upper(), ".NS", ".BO"):
                for r_quote in results:
                    sym = r_quote.get("symbol", "")
                    if sym.upper().endswith(preference):
                        _ticker_cache[cache_key] = sym
                        save_cache()
                        return sym
    except Exception as e:
        print(f"[Yahoo Search API] '{name_clean}': {e}")

    _ticker_cache[cache_key] = None
    save_cache()
    return None


def _resolve_mf_scheme_code(fund_name: str) -> int | None:
    """
    Search and resolve mutual fund name to its scheme code from api.mfapi.in.
    Results are cached in _mf_scheme_cache.
    """
    cache_key = fund_name.strip()
    if cache_key in _mf_scheme_cache:
        return _mf_scheme_cache[cache_key]

    try:
        url = f"https://api.mfapi.in/mf/search?q={requests.utils.quote(cache_key)}"
        r = requests.get(url, timeout=5)
        if r.status_code == 200:
            results = r.json()
            if results:
                code = int(results[0]['schemeCode'])
                _mf_scheme_cache[cache_key] = code
                save_cache()
                return code
    except Exception as e:
        print(f"[MF scheme search] {fund_name}: {e}")

    return None


def _fetch_mf_history(scheme_code: int) -> pd.Series:
    """
    Fetch historical NAVs from api.mfapi.in and return as a pandas Series indexed by date.
    Returns empty series on failure.
    """
    try:
        url = f"https://api.mfapi.in/mf/{scheme_code}"
        r = requests.get(url, timeout=5)
        if r.status_code == 200:
            data = r.json().get('data', [])
            dates = []
            navs = []
            for item in data:
                # date format is "DD-MM-YYYY"
                dt = parse_date(item['date'])
                if dt:
                    dates.append(dt)
                    navs.append(float(item['nav']))
            
            s = pd.Series(navs, index=dates)
            s = s.sort_index()
            return s
    except Exception as e:
        print(f"[MF history fetch] {scheme_code}: {e}")
    return pd.Series(dtype=float)


def _yf_info(symbol: str, bypass_cache: bool = False) -> dict:
    """
    Fetch price and sector for a resolved Yahoo Finance symbol via yfinance.
    Returns a dict with keys 'price' (float|None), 'sector' (str|None), and 'beta' (float|None).
    """
    now = time.time()
    
    # If not bypassing cache, check if we have a fresh cached entry with price and recent timestamp
    if not bypass_cache and symbol in _info_cache:
        cached = _info_cache[symbol]
        if cached.get("price_timestamp") and now - cached["price_timestamp"] < 300:
            return cached

    # Pre-populate sector & beta from cached entry if they exist to avoid making an expensive ticker.info call
    cached_sector = None
    cached_beta = None
    if symbol in _info_cache:
        cached_sector = _info_cache[symbol].get("sector")
        cached_beta = _info_cache[symbol].get("beta")

    result = {
        "price": None,
        "sector": cached_sector,
        "beta": cached_beta,
        "price_timestamp": None
    }
    
    try:
        ticker = yf.Ticker(symbol)

        # Price: fast_info is lightweight and fast
        try:
            price = ticker.fast_info.last_price
            if price:
                result["price"] = round(float(price), 2)
                result["price_timestamp"] = now
        except Exception:
            pass

        # Sector & Beta: from the full info dict (only if not already cached)
        if not result["sector"] or result["beta"] is None:
            try:
                info = ticker.info
                if not result["price"]:
                    mp = info.get("regularMarketPrice") or info.get("currentPrice")
                    if mp:
                        result["price"] = round(float(mp), 2)
                        result["price_timestamp"] = now
                raw_sector = info.get("sector", "")
                if raw_sector:
                    result["sector"] = _SECTOR_MAP.get(raw_sector, raw_sector)
                
                beta = info.get("beta")
                if beta is not None:
                    result["beta"] = round(float(beta), 3)
            except Exception:
                pass

    except Exception as e:
        print(f"[yfinance info] {symbol}: {e}")

    # Save to persistent cache if we resolved a price
    if result["price"] is not None:
        _info_cache[symbol] = result
        save_cache()
        
    return result



# Indian market holidays (NSE/BSE) — add years as needed
_INDIAN_MARKET_HOLIDAYS = {
    # 2024
    datetime(2024, 1, 26), datetime(2024, 3, 25), datetime(2024, 3, 29),
    datetime(2024, 4, 14), datetime(2024, 4, 17), datetime(2024, 5, 23),
    datetime(2024, 6, 17), datetime(2024, 7, 17), datetime(2024, 8, 15),
    datetime(2024, 10, 2), datetime(2024, 10, 14), datetime(2024, 10, 24),
    datetime(2024, 11, 1), datetime(2024, 11, 15), datetime(2024, 12, 25),
    # 2025
    datetime(2025, 1, 26), datetime(2025, 2, 26), datetime(2025, 3, 14),
    datetime(2025, 3, 31), datetime(2025, 4, 10), datetime(2025, 4, 14),
    datetime(2025, 4, 18), datetime(2025, 5, 1), datetime(2025, 8, 15),
    datetime(2025, 8, 27), datetime(2025, 10, 2), datetime(2025, 10, 2),
    datetime(2025, 10, 20), datetime(2025, 10, 24), datetime(2025, 11, 5),
    datetime(2025, 12, 25),
    # 2026
    datetime(2026, 1, 26), datetime(2026, 3, 20), datetime(2026, 3, 31),
    datetime(2026, 4, 3), datetime(2026, 4, 14), datetime(2026, 5, 1),
    datetime(2026, 7, 6), datetime(2026, 8, 15), datetime(2026, 9, 16),
    datetime(2026, 10, 2), datetime(2026, 10, 14), datetime(2026, 11, 11),
    datetime(2026, 12, 25),
}

def _count_trading_days(start_date: datetime, end_date: datetime) -> int:
    """Count Mon-Fri trading days between two dates, excluding Indian market holidays."""
    count = 0
    d = start_date
    # Normalize holiday set to date-only for comparison
    holidays_dates = {h.date() for h in _INDIAN_MARKET_HOLIDAYS}
    while d.date() <= end_date.date():
        if d.weekday() < 5 and d.date() not in holidays_dates:  # Mon=0 … Fri=4
            count += 1
        d = d.replace(day=d.day + 1) if d.day < 28 else datetime(d.year + (d.month == 12), (d.month % 12) + 1 if d.month < 12 else 1, 1) if d.day >= 28 and d == datetime(d.year, d.month, d.day) else d
    return count

def _add_one_day(d: datetime) -> datetime:
    import calendar
    year, month, day = d.year, d.month, d.day
    day += 1
    max_day = calendar.monthrange(year, month)[1]
    if day > max_day:
        day = 1
        month += 1
        if month > 12:
            month = 1
            year += 1
    return datetime(year, month, day)

def _count_trading_days_v2(start_date: datetime, end_date: datetime) -> int:
    """Count Mon-Fri trading days between start_date and end_date (inclusive), excluding Indian holidays."""
    count = 0
    holidays_dates = {h.date() for h in _INDIAN_MARKET_HOLIDAYS}
    d = datetime(start_date.year, start_date.month, start_date.day)
    end = datetime(end_date.year, end_date.month, end_date.day)
    while d <= end:
        if d.weekday() < 5 and d.date() not in holidays_dates:
            count += 1
        d = _add_one_day(d)
    return count


def get_days_in_drawdown(symbol: str, buy_price: float, buy_date_raw=None) -> int:
    """
    Calculate how many consecutive trading days the stock's closing price has been
    strictly below the average purchase price (buy_price), anchored at buy_date.

    Returns:
      >= 0  : trading days in drawdown
      -1    : buy_date missing, duration cannot be calculated
    """
    if not symbol or not buy_price:
        return 0

    cache_key = f"{symbol}|{buy_price}"
    now_ts = time.time()

    # Check cache (1 hour TTL)
    if cache_key in _drawdown_cache:
        cached = _drawdown_cache[cache_key]
        if now_ts - cached.get("ts", 0) < 3600:
            return cached.get("days", 0)

    # If no buy_date provided, return sentinel -1 (unknown duration)
    buy_date = parse_date(buy_date_raw) if buy_date_raw else None
    if buy_date is None:
        _drawdown_cache[cache_key] = {"days": -1, "ts": now_ts}
        save_cache()
        return -1

    days = 0
    try:
        ticker = yf.Ticker(symbol)
        # Fetch up to 2 years of daily data
        hist = ticker.history(period="2y")
        if not hist.empty and "Close" in hist.columns:
            # Filter to only rows on/after buy_date
            hist.index = hist.index.tz_localize(None) if hist.index.tzinfo is not None else hist.index
            hist_after_buy = hist[hist.index >= pd.Timestamp(buy_date)]

            # Walk backward from today: count consecutive days below buy_price
            closes = hist_after_buy["Close"].tolist()
            count = 0
            for price in reversed(closes):
                if price < buy_price:
                    count += 1
                else:
                    break
            days = count
    except Exception as e:
        print(f"[drawdown calculation] {symbol}: {e}")
        days = 0

    # Update cache
    _drawdown_cache[cache_key] = {"days": days, "ts": now_ts}
    save_cache()
    return days



def get_yahoo_finance_price(company_name: str, exchange: str) -> float | None:
    """Return live price for a company name / exchange pair, or None."""
    sym = _resolve_ticker(company_name, exchange)
    if not sym:
        return None
    return _yf_info(sym)["price"]


def get_yahoo_sector(company_name: str, exchange: str) -> str | None:
    """Return sector string for a company name / exchange pair, or None."""
    sym = _resolve_ticker(company_name, exchange)
    if not sym:
        return None
    return _yf_info(sym)["sector"]


# Load Supabase data with calculations and returns
def get_portfolio_data(user_id=None):
    if not user_id:
        return {"stocks": [], "mfs": [], "summary": {}, "signals": []}

    config = load_config(user_id)
    stock_ltcg_days = config.get("tax_rules", {}).get("stocks_ltcg_days", 365)
    mf_ltcg_days = config.get("tax_rules", {}).get("mf_ltcg_days", 365)
    
    stocks = []
    mfs = []
    
    from lib.supabase_data import get_user_stock_holdings, get_user_mf_holdings
    
    # 1. READ STOCKS
    db_stocks = get_user_stock_holdings(user_id)
    for s in db_stocks:
        scrip = s.get('scrip_name')
        exchange = s.get('exchange', 'NSE') or 'NSE'
        sym = _resolve_ticker(str(scrip).strip(), str(exchange).strip())
        
        raw_sector = s.get('sector')
        sector, industry, is_sector_missing = resolve_stock_sector(raw_sector, sym)
        
        qty = s.get('quantity') or 0
        buy_price = float(s.get('buy_price') or 0)
        buy_date_raw = s.get('buy_date')
        raw_cp = s.get('current_price')
        if raw_cp is not None and float(raw_cp) > 0:
            current_price = float(raw_cp)
        else:
            # current_price is NULL or 0 — fetch live and persist it
            live_price = None
            if sym:
                try:
                    live_price = _yf_info(sym, bypass_cache=True).get('price')
                except Exception:
                    pass
            if live_price and live_price > 0:
                current_price = live_price
                from lib.supabase_data import update_stock_prices
                update_stock_prices(user_id, [{'id': s.get('id'), 'current_price': live_price}])
            else:
                current_price = buy_price  # genuine fallback: no price available
        
        cagr_5y = None
        nifty_cagr_5y = None
        dividends = 0.0
        
        invested_value = qty * buy_price
        current_value = qty * current_price
        pnl = current_value - invested_value
        return_pct = (pnl / invested_value) if invested_value > 0 else 0
        tax_flag = get_tax_flag(buy_date_raw, stock_ltcg_days)
        total_return = pnl + dividends
        
        buy_date_str = format_date_str(buy_date_raw) if buy_date_raw else ""
        holding_yrs = get_holding_period_years(buy_date_raw) if buy_date_raw else None
        holding_type = "LT" if tax_flag == "LTCG" else ("ST" if tax_flag == "STCG" else "")
        
        drawdown_days = get_days_in_drawdown(sym, buy_price, buy_date_raw) if sym else 0

        stocks.append({
            "id": str(s.get('id')),
            "row_idx": str(s.get('id')),
            "Scrip Name": str(scrip).strip(),
            "Exchange": str(exchange).strip(),
            "Sector": sector,
            "Industry": industry,
            "is_sector_missing": is_sector_missing,
            "Qty": qty,
            "Buy Price": buy_price,
            "Buy Date": buy_date_str,
            "Holding Period Yrs": holding_yrs,
            "Holding Type": holding_type,
            "Current Price": current_price,
            "5Y CAGR": cagr_5y,
            "Nifty 5Y CAGR": nifty_cagr_5y,
            "Invested Value": round(invested_value, 2),
            "Current Value": round(current_value, 2),
            "P&L": round(pnl, 2),
            "Return %": round(return_pct * 100, 2),
            "Dividends": round(dividends, 2),
            "Tax Flag": tax_flag,
            "Total Return": round(total_return, 2),
            "Drawdown Days": drawdown_days
        })

    # 2. READ MUTUAL FUNDS
    db_mfs = get_user_mf_holdings(user_id)
    for m in db_mfs:
        fund_name = m.get('fund_name')
        amc = m.get('amc') or 'Other'
        category = m.get('category') or 'Other'
        sub_category = m.get('sub_category') or 'Other'
        units = float(m.get('units_held') or 0)
        buy_nav = float(m.get('buy_nav') or 0)
        buy_date_raw = m.get('purchase_date')
        current_nav = float(m.get('current_nav') or 0) if m.get('current_nav') is not None else buy_nav
        
        benchmark_return = None
        expense_ratio = 0.0
        
        invested_value = units * buy_nav
        current_value = units * current_nav
        pnl = current_value - invested_value
        return_pct = (pnl / invested_value) if invested_value > 0 else 0
        tax_flag = get_tax_flag(buy_date_raw, mf_ltcg_days)
        holding_yrs = get_holding_period_years(buy_date_raw) if buy_date_raw else 0
        holding_type = "LT" if tax_flag == "LTCG" else "ST"
        xirr_pct = compute_xirr(invested_value, current_value, buy_date_raw) if buy_date_raw else None
        estimated_tax = compute_mf_tax(pnl, tax_flag)
        investment_type = "Lumpsum"
        
        buy_date_str = format_date_str(buy_date_raw) if buy_date_raw else ""
        
        mfs.append({
            "id": str(m.get('id')),
            "row_idx": str(m.get('id')),
            "Fund Name": str(fund_name).strip(),
            "AMC": str(amc).strip(),
            "Category": str(category).strip(),
            "Sub-Category": str(sub_category).strip(),
            "Units Held": units,
            "Buy NAV": buy_nav,
            "Buy Date": buy_date_str,
            "Current NAV": current_nav,
            "Benchmark 1Y Return %": benchmark_return,
            "Expense Ratio %": expense_ratio,
            "Invested Value": round(invested_value, 2),
            "Current Value": round(current_value, 2),
            "P&L": round(pnl, 2),
            "Absolute Return %": round(return_pct * 100, 2),
            "Tax Flag": tax_flag,
            "XIRR %": xirr_pct,
            "Holding Period Yrs": holding_yrs,
            "Holding Type": holding_type,
            "Estimated Tax": estimated_tax,
            "Investment Type": str(investment_type).strip()
        })

    # Calculations for summary stats
    total_stock_invested = sum(s["Invested Value"] for s in stocks)
    total_stock_value = sum(s["Current Value"] for s in stocks)
    total_stock_pnl = total_stock_value - total_stock_invested
    total_stock_return_pct = (total_stock_pnl / total_stock_invested * 100) if total_stock_invested > 0 else 0
    total_stock_dividends = sum(s["Dividends"] for s in stocks)
    
    # Calculate weighted Portfolio Beta
    total_beta_weighted = 0.0
    for s in stocks:
        val = s["Current Value"]
        scrip = s["Scrip Name"]
        exchange = s["Exchange"]
        sym = _resolve_ticker(scrip, exchange)
        beta = 1.0  # default
        if sym and sym in _info_cache:
            info_beta = _info_cache[sym].get("beta")
            if isinstance(info_beta, (int, float)):
                beta = info_beta
        total_beta_weighted += beta * val
    portfolio_beta = (total_beta_weighted / total_stock_value) if total_stock_value > 0 else 1.0
    
    total_mf_invested = sum(m["Invested Value"] for m in mfs)
    total_mf_value = sum(m["Current Value"] for m in mfs)
    total_mf_pnl = total_mf_value - total_mf_invested
    total_mf_return_pct = (total_mf_pnl / total_mf_invested * 100) if total_mf_invested > 0 else 0
    total_mf_stcg_tax = sum(m["Estimated Tax"] for m in mfs if m["Tax Flag"] == "STCG")
    total_mf_ltcg_tax = sum(m["Estimated Tax"] for m in mfs if m["Tax Flag"] == "LTCG")
    
    total_portfolio_invested = total_stock_invested + total_mf_invested
    total_portfolio_value = total_stock_value + total_mf_value
    total_portfolio_pnl = total_portfolio_value - total_portfolio_invested
    total_portfolio_return_pct = (total_portfolio_pnl / total_portfolio_invested * 100) if total_portfolio_invested > 0 else 0
    
    # Risk signals: computed dynamically from Supabase holdings, merged per stock
    _PRIORITY_ORDER = {"CRITICAL": 3, "HIGH": 2, "LOW": 1}
    signals_by_stock: dict = {}
    total_stock_value_for_signals = sum(s["Current Value"] for s in stocks)

    for s in stocks:
        scrip      = s["Scrip Name"]
        cur_val    = s["Current Value"]
        ret_pct    = s["Return %"]
        weight_pct = (cur_val / total_stock_value_for_signals * 100) if total_stock_value_for_signals > 0 else 0

        raw_signals = []  # (signal_type, priority, action)

        # Critical Loss: return below -30%
        if ret_pct < -30:
            raw_signals.append(("Critical Loss", "CRITICAL", "Exit immediately or average down with caution"))
        # Underperformer: return between -30% and -10%
        elif ret_pct < -10:
            raw_signals.append(("Underperformer", "HIGH", "Review thesis and consider exit"))
        # Strong Performer: return above 20%
        elif ret_pct > 20:
            raw_signals.append(("Strong Performer", "LOW", "Consider booking partial profits"))

        # Concentration Risk: position weight above 20% of stock portfolio (independent check)
        if weight_pct > 20:
            raw_signals.append(("Concentration Risk", "HIGH",
                                f"Reduce position size (currently {weight_pct:.1f}% of portfolio)"))

        if not raw_signals:
            continue  # stock is within normal range — no signal needed

        if scrip not in signals_by_stock:
            signals_by_stock[scrip] = {
                "Scrip Name": scrip,
                "Current Value": cur_val,
                "Return %": ret_pct,
                "Signals": [],
                "Priority": "LOW",
                "Actions": []
            }

        for sig_type, sig_priority, sig_action in raw_signals:
            signals_by_stock[scrip]["Signals"].append(sig_type)
            signals_by_stock[scrip]["Actions"].append(sig_action)
            # Escalate to highest priority seen
            if _PRIORITY_ORDER.get(sig_priority, 0) > _PRIORITY_ORDER.get(signals_by_stock[scrip]["Priority"], 0):
                signals_by_stock[scrip]["Priority"] = sig_priority

    # Convert to final list — join multi-signals per stock into plain strings
    signals = []
    for entry in signals_by_stock.values():
        sig_str    = " + ".join(entry["Signals"])
        action_str = " • ".join(entry["Actions"])
        signals.append({
            "Scrip Name":         entry["Scrip Name"],
            "Current Value":      entry["Current Value"],
            "Return %":           entry["Return %"],
            "Signal":             sig_str,
            "Signals":            entry["Signals"],   # keep list for JS multi-badge
            "Priority":           entry["Priority"],
            "Recommended Action": action_str,
            "Actions":            entry["Actions"],   # keep list for JS join
        })

    return {
        "stocks": stocks,
        "mfs": mfs,
        "signals": signals,
        "summary": {
            "total_stock_invested": round(total_stock_invested, 2),
            "total_stock_value": round(total_stock_value, 2),
            "total_stock_pnl": round(total_stock_pnl, 2),
            "total_stock_return_pct": round(total_stock_return_pct, 2),
            "total_stock_dividends": round(total_stock_dividends, 2),
            "portfolio_beta": round(portfolio_beta, 2),
            
            "total_mf_invested": round(total_mf_invested, 2),
            "total_mf_value": round(total_mf_value, 2),
            "total_mf_pnl": round(total_mf_pnl, 2),
            "total_mf_return_pct": round(total_mf_return_pct, 2),
            "total_mf_stcg_tax": round(total_mf_stcg_tax, 2),
            "total_mf_ltcg_tax": round(total_mf_ltcg_tax, 2),
            
            "total_portfolio_invested": round(total_portfolio_invested, 2),
            "total_portfolio_value": round(total_portfolio_value, 2),
            "total_portfolio_pnl": round(total_portfolio_pnl, 2),
            "total_portfolio_return_pct": round(total_portfolio_return_pct, 2)
        }
    }

# Sync formulas and styles on headers to match formatting
def apply_excel_styles(ws, row):
    # Apply standard fonts and alignments to avoid ugly Excel layouts
    thin_font = Font(name="Calibri", size=11)
    align_center = Alignment(horizontal="center")
    align_left = Alignment(horizontal="left")
    align_right = Alignment(horizontal="right")
    
    for col in range(1, 25):
        cell = ws.cell(row=row, column=col)
        if cell.value is not None:
            cell.font = thin_font
            if col in [1, 2, 3, 6, 17]:  # Text & Dates
                cell.alignment = align_center if col != 1 else align_left
            else:  # Numbers
                cell.alignment = align_right

# Background thread for live price updates
def run_price_and_nav_update(user_id, is_background=False):
    global price_update_state
    
    if "users" not in price_update_state:
        price_update_state["users"] = {}
        
    user_state = price_update_state["users"].setdefault(user_id, {
        "status": "idle",
        "logs": [],
        "start_time": 0,
        "current_index": 0,
        "total_tickers": 0,
        "current_ticker": ""
    })
    
    # Avoid overlapping updates for this user
    if user_state["status"] in ("running", "background_running"):
        return False
        
    user_state["status"] = "background_running" if is_background else "running"
    user_state["logs"] = []
    user_state["start_time"] = time.time()
    user_state["current_index"] = 0
    user_state["total_tickers"] = 0
    user_state["current_ticker"] = ""
    
    try:
        from lib.supabase_data import get_user_stock_holdings, get_user_mf_holdings, update_stock_prices, update_mf_holding
        
        # 1. Read Stocks to update
        db_stocks = get_user_stock_holdings(user_id)
        tickers_to_update = []
        for s in db_stocks:
            scrip = s.get('scrip_name')
            if scrip:
                exchange = s.get('exchange', 'NSE') or 'NSE'
                tickers_to_update.append(("stock", s.get('id'), str(scrip).strip(), str(exchange).strip()))
                
        # 2. Read Mutual Funds to update
        db_mfs = get_user_mf_holdings(user_id)
        mf_to_update = []
        for m in db_mfs:
            fund_name = m.get('fund_name')
            if fund_name:
                mf_to_update.append(("mf", m.get('id'), str(fund_name).strip(), None))
                
        items_to_update = tickers_to_update + mf_to_update
        user_state["total_tickers"] = len(items_to_update)
        user_state["logs"].append(f"Starting update. Found {len(tickers_to_update)} stocks and {len(mf_to_update)} mutual funds to update.")
        
        updated_count = 0
        
        for idx, (item_type, row_id, name, exchange) in enumerate(items_to_update):
            user_state["current_index"] = idx + 1
            user_state["current_ticker"] = name
            
            if item_type == "stock":
                user_state["logs"].append(f"Fetching live price for stock {name} ({exchange})...")
                sym = _resolve_ticker(name, exchange)
                if not sym:
                    user_state["logs"].append(f"  -> ❌ Could not resolve ticker symbol.")
                    time.sleep(0.1)
                    continue
                    
                info = _yf_info(sym, bypass_cache=True)
                price = info["price"]
                sector = info["sector"]
                
                if price is not None:
                    update_stock_prices(user_id, [{'id': row_id, 'current_price': price}])
                    log_line = f"  -> ✅ ₹{price}  [{sym}]"
                    
                    # Auto-fill sector if missing
                    existing_stock = next((s for s in db_stocks if s.get('id') == row_id), None)
                    current_sector = str(existing_stock.get('sector') or "").strip() if existing_stock else ""
                    if current_sector in ("", "Other", "ETFs") and sector:
                        from lib.supabase_data import update_stock_holding
                        update_stock_holding(user_id, row_id, {'Sector': sector})
                        log_line += f"  | Sector: {sector}"
                    
                    user_state["logs"].append(log_line)
                    updated_count += 1
                else:
                    user_state["logs"].append(f"  -> ❌ Price not available for [{sym}].")
            else:
                user_state["logs"].append(f"Fetching live NAV for mutual fund {name}...")
                code = _resolve_mf_scheme_code(name)
                if not code:
                    user_state["logs"].append(f"  -> ❌ Could not resolve scheme code.")
                    time.sleep(0.1)
                    continue
                    
                try:
                    url = f"https://api.mfapi.in/mf/{code}"
                    r = requests.get(url, timeout=5)
                    if r.status_code == 200:
                        res_data = r.json().get('data', [])
                        if res_data:
                            nav = float(res_data[0]['nav'])
                            update_mf_holding(user_id, row_id, {'Current NAV': nav})
                            user_state["logs"].append(f"  -> ✅ NAV: ₹{nav}")
                            updated_count += 1
                        else:
                            user_state["logs"].append(f"  -> ❌ No NAV data found.")
                    else:
                        user_state["logs"].append(f"  -> ❌ API error ({r.status_code}).")
                except Exception as e:
                    user_state["logs"].append(f"  -> ❌ Connection error: {e}")
                    
            time.sleep(0.3)
            
        user_state["status"] = "completed"
        price_update_state["last_updated_by_profile"][user_id] = time.time()
        user_state["logs"].append(f"Successfully updated {updated_count} items and saved to Supabase.")
        return True
    except Exception as e:
        user_state["status"] = "error"
        user_state["last_failed_time"] = time.time()
        user_state["logs"].append(f"Critical Error during update: {e}")
        raise

# Background thread for live price updates (manual trigger wrapper)
def live_price_updater_thread(user_id):
    try:
        run_price_and_nav_update(user_id, is_background=False)
    except Exception:
        pass

# Lightweight helper: fetch+write prices for one user (used by /api/update-prices)
def run_price_update_for_user(user_id: str):
    """Fetch live prices for every stock holding of a user and write to Supabase."""
    from lib.supabase_data import get_user_stock_holdings, update_stock_prices as _usp
    try:
        holdings = get_user_stock_holdings(user_id)
        updates = []
        for h in holdings:
            scrip = h.get('scrip_name', '')
            exchange = h.get('exchange', 'NSE') or 'NSE'
            if not scrip:
                continue
            sym = _resolve_ticker(str(scrip).strip(), str(exchange).strip())
            if not sym:
                print(f"[PriceUpdate] Could not resolve ticker for: {scrip}")
                continue
            info = _yf_info(sym, bypass_cache=True)
            price = info.get('price')
            if price is not None:
                updates.append({'id': h['id'], 'current_price': price})
                print(f"[PriceUpdate] {scrip} ({sym}) -> ₹{price}")
            else:
                print(f"[PriceUpdate] No price for {scrip} ({sym})")
        if updates:
            _usp(user_id, updates)
            print(f"[PriceUpdate] Updated {len(updates)} stock prices for user {user_id}")
        else:
            print(f"[PriceUpdate] No prices fetched for user {user_id}")
    except Exception as e:
        print(f"[PriceUpdate] ERROR for user {user_id}: {e}")

# Background daemon loop for automatic updates
def automatic_price_updater_loop():
    print("[Auto Price Updater] Starting background automatic price updater loop...")
    time.sleep(15)  # wait for Flask app to fully spin up
    
    while True:
        try:
            from lib.supabase import get_supabase_admin
            supabase = get_supabase_admin()
            res = supabase.table('profiles').select('id').execute()
            user_ids = [item['id'] for item in res.data] if res.data else []
            
            for user_id in user_ids:
                user_config = load_config(user_id)
                if user_config.get("auto_update_prices", True):
                    last_updated = price_update_state["last_updated_by_profile"].get(user_id, 0)
                    now = time.time()
                    if now - last_updated >= 300: # 5 minutes
                        print(f"[Auto Price Updater] Running scheduled automatic price/NAV update for user: {user_id}")
                        try:
                            run_price_and_nav_update(user_id, is_background=True)
                        except Exception as e:
                            print(f"[Auto Price Updater] Scheduled update failed for user {user_id}: {e}")
        except Exception as e:
            print(f"[Auto Price Updater] Exception in background loop: {e}")
            
        time.sleep(60)  # Check every 60 seconds

@app.route('/')
def home():
    return render_template('index.html', **_auth_context())

# ─── Authentication Pages ──────────────────────────────────────────────────────
def _auth_context():
    """Return Supabase credentials for client-side auth templates."""
    return {
        'supabase_url': os.environ.get('SUPABASE_URL', ''),
        'supabase_anon_key': os.environ.get('SUPABASE_ANON_KEY', ''),
    }

@app.route('/signup')
def signup_page():
    return render_template('signup.html', **_auth_context())

@app.route('/login')
def login_page():
    return render_template('login.html', **_auth_context())

@app.route('/verify-email')
def verify_email_page():
    return render_template('verify-email.html', **_auth_context())

@app.route('/reset-password')
def reset_password_page():
    return render_template('reset-password.html', **_auth_context())

@app.route('/dashboard')
def dashboard_page():
    return render_template('index.html', **_auth_context())

@app.route('/account')
def account_page():
    return render_template('account.html', **_auth_context())

@app.route('/terms')
def terms_page():
    return render_template('terms.html')

@app.route('/privacy')
def privacy_page():
    return render_template('privacy.html')

@app.route('/disclaimer')
def disclaimer_page():
    return render_template('disclaimer.html')

@app.route('/api/account/profile', methods=['GET'])
def get_account_profile():
    user_id = get_current_user_id()
    if not user_id:
        return jsonify({"status": "error", "message": "Unauthorized"}), 401
    try:
        supabase = get_supabase_admin()
        result = supabase.table('profiles')\
            .select('*')\
            .eq('id', user_id)\
            .single()\
            .execute()
        profile = result.data
        return jsonify({
            "status": "success",
            "profile": {
                "id": profile.get('id'),
                "email": profile.get('email'),
                "full_name": profile.get('full_name', ''),
                "plan": profile.get('plan', 'free'),
                "plan_expires_at": profile.get('plan_expires_at', ''),
                "created_at": profile.get('created_at', '')
            }
        })
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/account/update-name', methods=['POST'])
def update_account_name():
    user_id = get_current_user_id()
    if not user_id:
        return jsonify({"status": "error", "message": "Unauthorized"}), 401
    data = request.json
    new_name = data.get('full_name', '').strip()
    if not new_name:
        return jsonify({"status": "error", "message": "Name cannot be empty"}), 400
    try:
        supabase = get_supabase_admin()
        supabase.table('profiles')\
            .update({'full_name': new_name})\
            .eq('id', user_id)\
            .execute()
        return jsonify({"status": "success", "message": "Name updated successfully!"})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/account/delete', methods=['POST'])
def delete_account():
    user_id = get_current_user_id()
    if not user_id:
        return jsonify({"status": "error", "message": "Unauthorized"}), 401
    data = request.json
    confirmation = data.get('confirmation', '')
    if confirmation != 'DELETE':
        return jsonify({"status": "error", "message": "Please type DELETE to confirm"}), 400
    try:
        supabase = get_supabase_admin()
        supabase.table('stock_holdings').delete().eq('user_id', user_id).execute()
        supabase.table('mf_holdings').delete().eq('user_id', user_id).execute()
        supabase.table('subscriptions').delete().eq('user_id', user_id).execute()
        supabase.table('profiles').delete().eq('id', user_id).execute()
        return jsonify({"status": "success", "message": "Account deleted"})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/update-prices', methods=['POST'])
def trigger_price_update():
    """Manual trigger: fetch live prices and write them to Supabase for this user."""
    user_id = get_current_user_id()
    if not user_id:
        return jsonify({"status": "error", "message": "Unauthorized"}), 401
    threading.Thread(
        target=run_price_update_for_user,
        args=(user_id,),
        daemon=True
    ).start()
    return jsonify({"status": "success", "message": "Price update started in background"})

@app.route('/api/test-db', methods=['GET'])
def test_db():
    try:
        supabase = get_supabase()
        supabase.table('profiles').select('id', count='exact').limit(1).execute()

        return jsonify({
            "status": "Connected successfully",
            "message": "Supabase is working"
        })
    except Exception as error:
        return jsonify({
            "status": "Connection failed",
            "error": str(error)
        }), 500

@app.route('/api/portfolio', methods=['GET'])
def api_portfolio():
    user_id = get_current_user_id()
    if not user_id:
        return jsonify({"status": "error", "message": "Unauthorized"}), 401
    data = get_portfolio_data(user_id)
    # Inject last_updated timestamp
    if isinstance(data, dict):
        data["last_updated"] = price_update_state["last_updated_by_profile"].get(user_id)
    return jsonify(data)

@app.route('/api/user/plan', methods=['GET'])
def get_user_plan_api():
    user_id = get_current_user_id()
    if not user_id:
        return jsonify({
            "status": "error",
            "message": "Unauthorized"
        }), 401
    
    plan = get_user_plan(user_id)
    limits = PLAN_LIMITS.get(plan, 
                             PLAN_LIMITS['free'])
    
    from lib.supabase_data import \
        get_user_stock_holdings, \
        get_user_mf_holdings
    
    current_stocks = len(
        get_user_stock_holdings(user_id))
    current_mf = len(
        get_user_mf_holdings(user_id))
    
    response = jsonify({
        "plan": plan,
        "limits": limits,
        "usage": {
            "stocks": current_stocks,
            "mf": current_mf
        }
    })
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    return response

@app.route('/api/payment/create-order', methods=['POST'])
def create_payment_order():
    user_id = get_current_user_id()
    if not user_id:
        return jsonify({
            "status": "error",
            "message": "Unauthorized"
        }), 401
    
    client, key_id, key_secret = get_razorpay_client()
    
    if not client or not key_id or not key_secret:
        # Mock mode
        return jsonify({
            "status": "success",
            "order_id": f"order_mock_{user_id[:8]}",
            "amount": 19900,
            "currency": "INR",
            "key_id": "rzp_test_mock_key",
            "is_mock": True
        })
    
    try:
        # Create Razorpay order
        # Amount in paise (₹199 = 19900)
        order_data = {
            "amount": 19900,
            "currency": "INR",
            "receipt": f"order_{user_id[:8]}",
            "notes": {
                "user_id": user_id,
                "plan": "pro",
                "duration": "monthly"
            }
        }
        
        order = client.order.create(data=order_data)
        
        return jsonify({
            "status": "success",
            "order_id": order['id'],
            "amount": order['amount'],
            "currency": order['currency'],
            "key_id": key_id,
            "is_mock": False
        })
    except Exception as e:
        return jsonify({
            "status": "error",
            "message": str(e)
        }), 500

@app.route('/api/payment/verify', methods=['POST'])
def verify_payment():
    user_id = get_current_user_id()
    if not user_id:
        return jsonify({
            "status": "error",
            "message": "Unauthorized"
        }), 401
    
    data = request.json
    is_mock = data.get('is_mock', False)
    
    try:
        if is_mock:
            payment_id = data.get('razorpay_payment_id', 'pay_mock_default')
        else:
            client, key_id, key_secret = get_razorpay_client()
            if not client:
                raise ValueError("Razorpay client is not configured, but a real signature verification was requested.")
            
            # Verify payment signature
            params_dict = {
                'razorpay_order_id': data.get('razorpay_order_id'),
                'razorpay_payment_id': data.get('razorpay_payment_id'),
                'razorpay_signature': data.get('razorpay_signature')
            }
            
            client.utility.verify_payment_signature(params_dict)
            payment_id = data.get('razorpay_payment_id')
        
        # Payment verified — upgrade user using admin client (to bypass RLS)
        supabase = get_supabase_admin()
        
        # Calculate expiry (30 days)
        from datetime import timedelta
        expiry = datetime.now() + timedelta(days=30)
        
        # Update user plan to pro
        supabase.table('profiles')\
            .update({
                'plan': 'pro',
                'plan_expires_at': expiry.isoformat()
            })\
            .eq('id', user_id)\
            .execute()
        
        # Save subscription record
        supabase.table('subscriptions')\
            .insert({
                'user_id': user_id,
                'razorpay_payment_id': payment_id,
                'plan': 'pro',
                'status': 'active',
                'started_at': datetime.now().isoformat(),
                'expires_at': expiry.isoformat()
            })\
            .execute()
        
        return jsonify({
            "status": "success",
            "message": "Payment verified! Welcome to Pro plan.",
            "plan": "pro",
            "expires_at": expiry.isoformat()
        })
        
    except Exception as e:
        return jsonify({
            "status": "error",
            "message": "Payment verification failed: " + str(e)
        }), 400

@app.route('/api/payment/status', methods=['GET'])
def payment_status():
    user_id = get_current_user_id()
    if not user_id:
        return jsonify({
            "status": "error",
            "message": "Unauthorized"
        }), 401
    
    try:
        # Query plan status using admin client
        supabase = get_supabase_admin()
        result = supabase.table('profiles')\
            .select('plan, plan_expires_at')\
            .eq('id', user_id)\
            .single()\
            .execute()
        
        if not result.data:
            return jsonify({
                "plan": "free",
                "expires_at": None,
                "is_active": False
            })
            
        profile = result.data
        plan = profile.get('plan', 'free')
        expires_at = profile.get('plan_expires_at')
        
        # Check if plan has expired
        if expires_at and plan == 'pro':
            from datetime import timezone
            expiry = datetime.fromisoformat(expires_at.replace('Z', '+00:00'))
            if datetime.now(timezone.utc) > expiry:
                # Downgrade to free
                supabase.table('profiles')\
                    .update({'plan': 'free'})\
                    .eq('id', user_id)\
                    .execute()
                plan = 'free'
        
        return jsonify({
            "plan": plan,
            "expires_at": expires_at,
            "is_active": plan == 'pro'
        })
    except Exception as e:
        return jsonify({
            "status": "error",
            "message": str(e)
        }), 500

@app.route('/upgrade')
def upgrade_page():
    _, key_id, _ = get_razorpay_client()
    return render_template(
        'upgrade.html',
        razorpay_key_id=key_id,
        **_auth_context()
    )


@app.route('/api/admin/update-plan', methods=['POST'])
def admin_update_plan():
    # Simple admin endpoint to manually
    # upgrade users during testing
    # (will be replaced by Razorpay 
    #  webhook in production)
    
    admin_key = request.headers.get(
        'X-Admin-Key', '')
    if admin_key != os.environ.get(
        'ADMIN_SECRET_KEY', 'changeme'):
        return jsonify({
            "status": "error",
            "message": "Unauthorized"
        }), 401
    
    data = request.json
    user_email = data.get('email')
    new_plan = data.get('plan', 'pro')
    
    try:
        supabase = get_supabase()
        
        # Get user id from email
        result = supabase.table('profiles')\
            .update({'plan': new_plan})\
            .eq('email', user_email)\
            .execute()
        
        return jsonify({
            "status": "success",
            "message": f"Plan updated to {new_plan} for {user_email}"
        })
    except Exception as e:
        return jsonify({
            "status": "error",
            "message": str(e)
        }), 500



# Simple TTL cache for performance and sector contribution data
_perf_cache = {}   # key -> (timestamp, payload)
_PERF_TTL   = 600  # 10-minute cache

_contrib_cache = {}  # key -> (timestamp, payload)
_CONTRIB_TTL = 600

def clear_user_caches(user_id):
    if not user_id:
        return
    perf_keys = [k for k in _perf_cache.keys() if str(k).endswith(f"|{user_id}")]
    for k in perf_keys:
        _perf_cache.pop(k, None)
    contrib_keys = [k for k in _contrib_cache.keys() if str(k).endswith(f"|{user_id}")]
    for k in contrib_keys:
        _contrib_cache.pop(k, None)

@app.route('/api/portfolio/performance', methods=['GET'])
def portfolio_performance():
    user_id = get_current_user_id()
    if not user_id:
        return jsonify({"error": "Unauthorized"}), 401

    from lib.supabase_data import get_user_stock_holdings, get_user_mf_holdings
    stock_holdings_db = get_user_stock_holdings(user_id)
    if not stock_holdings_db:
        return jsonify({"error": "No holdings found"}), 404

    period    = request.args.get('period',    '1M')
    benchmark = request.args.get('benchmark', 'nifty50')

    cache_key = f"{period}|{benchmark}|{user_id}"
    if cache_key in _perf_cache:
        ts, payload = _perf_cache[cache_key]
        if time.time() - ts < _PERF_TTL:
            return jsonify(payload)

    PERIOD_MAP = {
        '1W':  ('7d',  '1d'),
        '1M':  ('1mo', '1d'),
        '3M':  ('3mo', '1d'),
        '1Y':  ('1y',  '1d'),
        'All': ('5y',  '1wk'),
    }
    yf_period, yf_interval = PERIOD_MAP.get(period, ('1mo', '1d'))

    BENCH_MAP = {
        'nifty50':       ('^NSEI',      'Nifty 50'),
        'sensex':        ('^BSESN',     'BSE Sensex'),
        'bse500':        ('BSE-500.BO', 'BSE 500'),
        'crisil_hybrid': ('^NSEI',      'CRISIL Hybrid Index'),
    }
    bench_sym, bench_name = BENCH_MAP.get(benchmark, ('^NSEI', 'Nifty 50'))

    # Read current holdings from Supabase
    try:
        from lib.supabase_data import get_user_stock_holdings, get_user_mf_holdings
        
        # 1. Read Stocks
        db_stocks = stock_holdings_db
        stock_holdings = []
        for s in db_stocks:
            scrip = s.get('scrip_name')
            exchange = s.get('exchange', 'NSE') or 'NSE'
            qty = s.get('quantity') or 0
            if float(qty) > 0:
                stock_holdings.append({'name': str(scrip).strip(),
                                       'exchange': str(exchange).strip(),
                                       'qty': float(qty)})
        
        # 2. Read Mutual Funds
        db_mfs = get_user_mf_holdings(user_id)
        mf_holdings = []
        for m in db_mfs:
            fund_name = m.get('fund_name')
            units = m.get('units_held') or 0
            if float(units) > 0:
                mf_holdings.append({'name': str(fund_name).strip(),
                                    'qty': float(units)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

    if not stock_holdings and not mf_holdings:
        return jsonify({'error': 'No holdings found'}), 404

    # Resolve stock tickers
    valid_stocks = []
    for h in stock_holdings:
        sym = _resolve_ticker(h['name'], h['exchange'])
        if sym:
            h['symbol'] = sym
            valid_stocks.append(h)

    stock_syms = [h['symbol'] for h in valid_stocks]
    all_syms   = stock_syms + [bench_sym]

    # Batch-download historical data
    try:
        raw = yf.download(
            tickers   = all_syms,
            period    = yf_period,
            interval  = yf_interval,
            auto_adjust = True,
            progress  = False,
            threads   = True,
        )
        if raw.empty:
            return jsonify({'error': 'No historical data returned'}), 404

        # Extract Close prices; handle both single and multi-ticker structures
        if isinstance(raw.columns, pd.MultiIndex):
            closes = raw['Close'].copy()
        else:
            closes = raw[['Close']].rename(columns={'Close': all_syms[0]})

        closes = closes.ffill().dropna(how='all')
    except Exception as e:
        return jsonify({'error': f'History download failed: {e}'}), 500

    # Build stock portfolio value series
    stock_series = pd.Series(0.0, index=closes.index)
    for h in valid_stocks:
        sym = h['symbol']
        if sym in closes.columns:
            stock_series += closes[sym].fillna(0) * h['qty']

    # Build mutual fund portfolio value series
    mf_series = pd.Series(0.0, index=closes.index)
    for h in mf_holdings:
        code = _resolve_mf_scheme_code(h['name'])
        if code:
            nav_series = _fetch_mf_history(code)
            if not nav_series.empty:
                # Align and fill missing dates
                aligned_nav = nav_series.reindex(closes.index).ffill().bfill()
                mf_series += aligned_nav * h['qty']

    # Check if we have valid MF data
    has_mf = not mf_series.empty and float(mf_series.iloc[-1]) > 0

    # Calculate combined portfolio series
    if has_mf:
        combined_series = stock_series + mf_series

    # Align benchmark to same dates
    bench_series = closes[bench_sym] if bench_sym in closes.columns else pd.Series(dtype=float)
    bench_series = bench_series.reindex(closes.index).ffill()

    def normalize(s):
        first = s.dropna().iloc[0] if not s.dropna().empty else 1.0
        return ((s / first) * 100).round(3)

    stock_norm  = normalize(stock_series)
    bench_norm = normalize(bench_series)

    # Post-process for CRISIL Hybrid Index
    if benchmark == 'crisil_hybrid':
        first_bench = bench_series.dropna().iloc[0] if not bench_series.dropna().empty else 1.0
        nifty_norm = ((bench_series / first_bench) * 100)
        
        # Simulated debt component at 6.5% CAGR
        start_date = nifty_norm.index[0]
        days = (nifty_norm.index - start_date).days
        debt_norm = 100.0 * (1.065 ** (days / 365.25))
        
        # Hybrid is 65% Equity (Nifty 50) and 35% Debt
        bench_norm = (0.65 * nifty_norm + 0.35 * debt_norm).round(3)

    if has_mf:
        combined_norm = normalize(combined_series)
        cur_port  = float(combined_norm.dropna().iloc[-1]) if not combined_norm.dropna().empty else 100.0
    else:
        cur_port  = float(stock_norm.dropna().iloc[-1])  if not stock_norm.dropna().empty  else 100.0

    cur_bench = float(bench_norm.dropna().iloc[-1]) if not bench_norm.dropna().empty else 100.0
    outperf   = round(cur_port - cur_bench, 2)

    dates     = [d.strftime('%Y-%m-%d') for d in stock_norm.index]
    stock_vals = [v if pd.notna(v) else None for v in stock_norm.tolist()]
    bench_vals= [v if pd.notna(v) else None for v in bench_norm.tolist()]

    payload = {
        'dates':            dates,
        'portfolio':        stock_vals,
        'benchmark':        bench_vals,
        'benchmark_name':   bench_name,
        'outperformance':   outperf,
        'period':           period,
    }

    if has_mf:
        payload['combined'] = [v if pd.notna(v) else None for v in combined_norm.tolist()]

    _perf_cache[cache_key] = (time.time(), payload)
    return jsonify(payload)

@app.route('/api/portfolio/sector-contribution', methods=['GET'])
def sector_contribution():
    user_id = get_current_user_id()
    if not user_id:
        return jsonify({"error": "Unauthorized"}), 401

    from lib.supabase_data import get_user_stock_holdings
    stock_holdings_db = get_user_stock_holdings(user_id)
    if not stock_holdings_db:
        return jsonify({"error": "No holdings found"}), 404

    period    = request.args.get('period',    '1M')
    benchmark = request.args.get('benchmark', 'nifty50')

    cache_key = f"{period}|{benchmark}|{user_id}"
    if cache_key in _contrib_cache:
        ts, payload = _contrib_cache[cache_key]
        if time.time() - ts < _CONTRIB_TTL:
            return jsonify(payload)

    PERIOD_MAP = {
        '1W':  ('7d',  '1d'),
        '1M':  ('1mo', '1d'),
        '3M':  ('3mo', '1d'),
        '1Y':  ('1y',  '1d'),
        'All': ('5y',  '1wk'),
    }
    yf_period, yf_interval = PERIOD_MAP.get(period, ('1mo', '1d'))

    BENCH_MAP = {
        'nifty50':  ('^NSEI',  'Nifty 50'),
        'sensex':   ('^BSESN', 'BSE Sensex'),
    }
    bench_sym, bench_name = BENCH_MAP.get(benchmark, ('^NSEI', 'Nifty 50'))

    try:
        from lib.supabase_data import get_user_stock_holdings
        db_stocks = stock_holdings_db
        holdings = []
        for s in db_stocks:
            scrip = s.get('scrip_name')
            exchange = s.get('exchange', 'NSE') or 'NSE'
            qty = s.get('quantity') or 0
            raw_sector = s.get('sector')
            sym = _resolve_ticker(str(scrip).strip(), str(exchange).strip())
            sector, industry, is_sector_missing = resolve_stock_sector(raw_sector, sym)
            if float(qty) > 0:
                holdings.append({
                    'name': str(scrip).strip(),
                    'exchange': str(exchange).strip(),
                    'sector': sector,
                    'qty': float(qty),
                    'symbol': sym
                })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

    if not holdings:
        return jsonify({'error': 'No holdings found'}), 404

    # Filter valid resolved tickers
    valid = [h for h in holdings if h.get('symbol')]

    if not valid:
        return jsonify({'error': 'Could not resolve any ticker symbols'}), 404

    stock_syms = [h['symbol'] for h in valid]
    all_syms   = stock_syms + [bench_sym]

    try:
        raw = yf.download(
            tickers   = all_syms,
            period    = yf_period,
            interval  = yf_interval,
            auto_adjust = True,
            progress  = False,
            threads   = True,
        )
        if raw.empty:
            return jsonify({'error': 'No historical data returned'}), 404

        if isinstance(raw.columns, pd.MultiIndex):
            closes = raw['Close'].copy()
        else:
            closes = raw[['Close']].rename(columns={'Close': all_syms[0]})

        closes = closes.ffill().dropna(how='all')
    except Exception as e:
        return jsonify({'error': f'History download failed: {e}'}), 500

    if closes.empty:
        return jsonify({'error': 'No close prices available'}), 404

    # Compute individual stock gain/loss
    stocks_details = []
    total_start_val = 0.0
    total_gain_rupees = 0.0

    for h in valid:
        sym = h['symbol']
        if sym in closes.columns and not closes[sym].dropna().empty:
            sym_prices = closes[sym].dropna()
            p_start = float(sym_prices.iloc[0])
            p_end = float(sym_prices.iloc[-1])
        else:
            p_start = p_end = 0.0

        start_val = p_start * h['qty']
        end_val = p_end * h['qty']
        gain_rupees = (p_end - p_start) * h['qty']

        total_start_val += start_val
        total_gain_rupees += gain_rupees

        stocks_details.append({
            'name': h['name'],
            'sector': h['sector'],
            'qty': h['qty'],
            'p_start': p_start,
            'p_end': p_end,
            'start_val': start_val,
            'end_val': end_val,
            'gain_rupees': gain_rupees
        })

    if total_start_val == 0:
        total_start_val = 1.0  # Avoid division by zero

    # Sector aggregation
    sectors_data = {}
    for sd in stocks_details:
        sec = sd['sector']
        if sec not in sectors_data:
            sectors_data[sec] = {
                'sector': sec,
                'start_val': 0.0,
                'end_val': 0.0,
                'gain_rupees': 0.0,
                'stocks': []
            }
        sectors_data[sec]['start_val'] += sd['start_val']
        sectors_data[sec]['end_val'] += sd['end_val']
        sectors_data[sec]['gain_rupees'] += sd['gain_rupees']
        sectors_data[sec]['stocks'].append(sd)

    # Convert to list and calculate weights/contributions
    sectors_list = []
    for sec, sdata in sectors_data.items():
        start_val = sdata['start_val']
        gain_rupees = sdata['gain_rupees']
        
        weight = (start_val / total_start_val * 100)
        contrib_pct = (gain_rupees / total_start_val * 100)
        return_pct = (gain_rupees / start_val * 100) if start_val > 0 else 0.0

        sec_stocks = []
        for st in sdata['stocks']:
            st_weight = (st['start_val'] / total_start_val * 100)
            st_return = (st['gain_rupees'] / st['start_val'] * 100) if st['start_val'] > 0 else 0.0
            st_contrib = (st['gain_rupees'] / total_start_val * 100)
            sec_stocks.append({
                'name': st['name'],
                'return_pct': round(st_return, 2),
                'contrib_pct': round(st_contrib, 2),
                'contrib_rupees': round(st['gain_rupees'], 2)
            })

        sectors_list.append({
            'sector': sec,
            'weight_pct': round(weight, 2),
            'return_pct': round(return_pct, 2),
            'contrib_pct': round(contrib_pct, 2),
            'contrib_rupees': round(gain_rupees, 2),
            'stocks': sec_stocks
        })

    # Benchmark calculation
    bench_return_pct = 0.0
    if bench_sym in closes.columns and not closes[bench_sym].dropna().empty:
        b_prices = closes[bench_sym].dropna()
        b_start = float(b_prices.iloc[0])
        b_end = float(b_prices.iloc[-1])
        if b_start > 0:
            bench_return_pct = ((b_end - b_start) / b_start * 100)

    portfolio_return_pct = (total_gain_rupees / total_start_val * 100)
    alpha = portfolio_return_pct - bench_return_pct

    payload = {
        'sectors': sectors_list,
        'portfolio_return_pct': round(portfolio_return_pct, 2),
        'benchmark_return_pct': round(bench_return_pct, 2),
        'alpha': round(alpha, 2),
        'benchmark_name': bench_name
    }

    _contrib_cache[cache_key] = (time.time(), payload)
    return jsonify(payload)


@app.route('/api/portfolio/reset', methods=['POST'])
def reset_portfolio():
    try:
        if not os.path.exists(EXCEL_FILE):
            return jsonify({"status": "error", "message": f"Excel file '{EXCEL_FILE}' not found."}), 404
            
        wb = openpyxl.load_workbook(EXCEL_FILE)
        
        # Wiping all stock holdings rows 4 to 1000
        if "📥 Stock Data Input" in wb.sheetnames:
            ws_s = wb["📥 Stock Data Input"]
            ws_s.delete_rows(4, 1000)
            
        # Wiping all Mutual Fund holdings rows 4 to 1000
        if "📥 MF Data Input" in wb.sheetnames:
            ws_m = wb["📥 MF Data Input"]
            ws_m.delete_rows(4, 1000)
            
        wb.save(EXCEL_FILE)
        return jsonify({
            "status": "success", 
            "message": f"All holdings for profile '{ACTIVE_PROFILE}' have been cleared successfully!"
        })
    except PermissionError:
        return jsonify({
            "status": "error", 
            "message": f"The Excel workbook '{EXCEL_FILE}' is currently locked by another program (e.g. MS Excel). Please close it and try again!"
        }), 423
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/config', methods=['GET', 'POST'])
def api_config():
    user_id = get_current_user_id()
    if not user_id:
        return jsonify({"status": "error", "message": "Unauthorized"}), 401
        
    if request.method == 'GET':
        return jsonify(load_config(user_id))
    else:
        new_config = request.json
        save_config(new_config, user_id)
        return jsonify({"status": "success", "config": new_config})

@app.route('/api/stock/live-price', methods=['GET'])
def get_live_stock_price():
    scrip = request.args.get('scrip', '').strip()
    if not scrip:
        return jsonify({"status": "error", "message": "Scrip name is required."}), 400
    
    price = get_yahoo_finance_price(scrip, "NSE")
    if price is not None:
        return jsonify({"status": "success", "price": price})
    else:
        return jsonify({"status": "error", "message": "Could not fetch live price."}), 404

@app.route('/api/stock/add', methods=['POST'])
def add_stock():
    user_id = get_current_user_id()
    if not user_id:
        return jsonify({"status": "error", "message": "Unauthorized"}), 401
        
    limit_check = check_stock_limit(user_id)
    if not limit_check["allowed"]:
        return jsonify({
            "status": "error",
            "message": f"Free plan limit reached. You have {limit_check['current']}/{limit_check['max']} stocks. Upgrade to Pro for unlimited stocks.",
            "upgrade_required": True,
            "feature": "max_stocks"
        }), 403
        
    data = request.json
    try:
        from lib.supabase_data import add_stock_holding
        company_name = data.get("Company Name", "").strip()
        sector       = data.get("Sector", "Other").strip() or "Other"
        qty          = float(data.get("Total Quantity", 0) or 0)
        avg_price    = float(data.get("Avg Trading Price", 0) or 0)

        # Auto-detect sector from Yahoo Finance when not provided
        if sector == "Other" and company_name:
            fetched_sector = get_yahoo_sector(company_name, "NSE")
            if fetched_sector:
                sector = fetched_sector

        # Resolve Current Price
        current_price = data.get("Current Price")
        if current_price is not None and str(current_price).strip() != "":
            try:
                current_price = float(current_price)
            except ValueError:
                current_price = None
        else:
            current_price = None

        if current_price is None:
            live_price = get_yahoo_finance_price(company_name, "NSE")
            current_price = live_price if live_price is not None else avg_price

        buy_date_raw = data.get("Buy Date")
        buy_date = parse_date(buy_date_raw) if buy_date_raw else None
        
        sym = _resolve_ticker(company_name, "NSE")
        sector, industry, is_sector_missing = resolve_stock_sector(sector, sym)

        holding_data = {
            'Scrip Name': company_name,
            'Exchange': 'NSE',
            'Sector': sector,
            'Industry': industry,
            'Qty': qty,
            'Buy Price': avg_price,
            'Buy Date': buy_date.strftime('%Y-%m-%d') if buy_date else None,
            'Current Price': current_price
        }

        res = add_stock_holding(user_id, holding_data)
        if res:
            clear_user_caches(user_id)
            return jsonify({"status": "success", "message": f"Stock '{company_name}' added successfully to Supabase!"})
        else:
            return jsonify({"status": "error", "message": "Failed to add stock to Supabase."}), 500
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/api/stock/import-csv', methods=['POST'])
def import_stocks_csv():
    user_id = get_current_user_id()
    if not user_id:
        return jsonify({"status": "error", "message": "Unauthorized"}), 401

    if 'file' not in request.files:
        return jsonify({"status": "error", "message": "No file uploaded. Please attach a CSV file."}), 400

    uploaded_file = request.files['file']
    if not uploaded_file.filename.lower().endswith('.csv'):
        return jsonify({"status": "error", "message": "Invalid file type. Please upload a .csv file."}), 400

    raw_bytes = uploaded_file.read()
    content = None
    used_encoding = None
    for enc in ('utf-8-sig', 'utf-8', 'cp1252', 'latin-1', 'iso-8859-1'):
        try:
            content = raw_bytes.decode(enc)
            used_encoding = enc
            break
        except (UnicodeDecodeError, LookupError):
            continue

    if content is None:
        return jsonify({"status": "error", "message": "Could not decode the CSV file. Please save it as UTF-8 and try again."}), 400

    try:
        reader = csv.reader(io.StringIO(content, newline=''))
        rows = list(reader)
    except Exception as e:
        return jsonify({"status": "error", "message": f"Failed to parse CSV: {e}"}), 400

    SYNONYMS = {
        'Company Name': {
            'company name', 'stock name', 'scrip name', 'security name',
            'script name', 'name', 'symbol', 'instrument', 'stock', 'scrip',
            'security', 'share name', 'script', 'isin name', 'isin_name'
        },
        'Total Quantity': {
            'total quantity', 'quantity', 'qty', 'net qty', 'total qty',
            'holding qty', 'shares', 'units', 'net quantity', 'holding quantity',
            'balance qty', 'available qty', 'free quantity', 'current free', 'current_free'
        },
        'Avg Trading Price': {
            'avg trading price', 'average buy price', 'avg buy price',
            'average price', 'avg price', 'avg. price', 'avg cost',
            'average cost', 'avg. cost', 'buy price', 'cost price',
            'avg cost price', 'purchase price', 'average purchase price',
            'avg purchase price', 'weighted avg price', 'rate'
        },
        'Sector': {
            'sector', 'industry', 'category', 'segment'
        }
    }

    header_row_idx = None
    col_map = {}

    for i, row in enumerate(rows):
        norm = [cell.strip().lower() for cell in row]
        matched = {}
        for logical_field, synonyms in SYNONYMS.items():
            for idx, cell in enumerate(norm):
                if cell in synonyms:
                    matched[logical_field] = idx
                    break

        if all(f in matched for f in ('Company Name', 'Total Quantity', 'Avg Trading Price')):
            header_row_idx = i
            col_map = matched
            break

    if header_row_idx is None:
        preview_lines = []
        for i, row in enumerate(rows[:20]):
            first_cells = ", ".join(f'"{c.strip()}"' for c in row[:4] if c.strip())
            if first_cells:
                preview_lines.append(f"Row {i+1}: [{first_cells}]")
        preview_str = " | ".join(preview_lines[:10]) or "File appears empty."
        return jsonify({
            "status": "error",
            "message": (
                f"Could not locate the equity holdings header row (encoding: {used_encoding}). "
                f"Looked for columns like 'Company Name'/'Stock Name', 'Quantity'/'Total Quantity', "
                f"and 'Avg Trading Price'/'Average buy price'. "
                f"First rows in file: {preview_str}"
            )
        }), 400

    STOP_KEYWORDS = {'total', 'grand total', 'sub total', 'subtotal', 'summary', ''}

    imported_stocks = []
    company_col = col_map['Company Name']
    qty_col     = col_map['Total Quantity']
    price_col   = col_map['Avg Trading Price']
    sector_col  = col_map.get('Sector')

    for row in rows[header_row_idx + 1:]:
        if not row or not any(cell.strip() for cell in row):
            break

        company_cell = row[company_col].strip() if company_col < len(row) else ''
        if '#' in company_cell:
            company_cell = company_cell.split('#')[0].strip()

        if company_cell.lower() in STOP_KEYWORDS:
            break

        if not company_cell:
            continue

        try:
            qty_raw   = row[qty_col].strip().replace(',', '') if qty_col < len(row) else ''
            qty       = float(qty_raw) if qty_raw else 0.0
            price_raw = row[price_col].strip().replace(',', '') if price_col < len(row) else ''
            avg_price = float(price_raw) if price_raw else 0.0

            sector = ''
            if sector_col is not None and sector_col < len(row):
                sector = row[sector_col].strip()

            imported_stocks.append({
                'Company Name':      company_cell,
                'Sector':            sector or 'Other',
                'Total Quantity':    qty,
                'Avg Trading Price': avg_price
            })
        except (ValueError, IndexError):
            continue

    if not imported_stocks:
        return jsonify({"status": "error", "message": "No valid stock data found in the uploaded CSV."}), 400

    try:
        from lib.supabase_data import add_stock_holding
        added = 0
        skipped = 0
        for stock in imported_stocks:
            company   = stock["Company Name"]
            sector    = stock["Sector"]
            qty       = stock["Total Quantity"]
            avg_price = stock["Avg Trading Price"]

            if qty <= 0:
                skipped += 1
                continue

            sym = _resolve_ticker(company, "NSE")
            sector, industry, is_sector_missing = resolve_stock_sector(sector, sym)

            holding_data = {
                'Scrip Name': company,
                'Exchange': 'NSE',
                'Sector': sector,
                'Industry': industry,
                'Qty': qty,
                'Buy Price': avg_price,
                'Buy Date': None,
                'Current Price': avg_price
            }
            res = add_stock_holding(user_id, holding_data)
            if res:
                added += 1
            else:
                skipped += 1

        if added > 0:
            clear_user_caches(user_id)

        msg = f"Successfully imported {added} stock(s) from CSV to Supabase!"
        if skipped > 0:
            msg += f" ({skipped} rows skipped — zero quantity or write error.)"
        return jsonify({"status": "success", "message": msg, "imported": added, "skipped": skipped})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/api/stock/edit', methods=['POST'])
def edit_stock():
    user_id = get_current_user_id()
    if not user_id:
        return jsonify({"status": "error", "message": "Unauthorized"}), 401
    data = request.json
    row_idx = (data.get("row_idx") or
               data.get("holding_id") or
               data.get("id"))
    try:
        from lib.supabase_data import update_stock_holding
        company_name = (data.get("Company Name") or data.get("Scrip Name", "")).strip()
        sector       = (data.get("Sector", "Other") or "Other").strip()
        qty          = float(data.get("Total Quantity") or data.get("Qty", 0) or 0)
        avg_price    = float(data.get("Avg Trading Price") or data.get("Buy Price", 0) or 0)

        # Resolve Current Price
        current_price = data.get("Current Price")
        if current_price is not None and str(current_price).strip() != "":
            try:
                current_price = float(current_price)
            except ValueError:
                current_price = None
        else:
            current_price = None

        if current_price is None:
            from lib.supabase_data import get_user_stock_holdings
            holdings = get_user_stock_holdings(user_id)
            existing = next((h for h in holdings if str(h['id']) == row_idx), None)
            existing_price = float(existing['current_price']) if existing and existing.get('current_price') is not None else None
            
            if existing and existing.get('scrip_name', '').upper() != company_name.upper() or existing_price is None or existing_price == 0:
                live_price = get_yahoo_finance_price(company_name, "NSE")
                current_price = live_price if live_price is not None else avg_price
            else:
                current_price = existing_price

        buy_date_raw = data.get("Buy Date")
        buy_date = parse_date(buy_date_raw) if buy_date_raw else None
        
        sym = _resolve_ticker(company_name, "NSE")
        sector, industry, is_sector_missing = resolve_stock_sector(sector, sym)

        holding_data = {
            'Scrip Name': company_name,
            'Exchange': 'NSE',
            'Sector': sector,
            'Industry': industry,
            'Qty': qty,
            'Buy Price': avg_price,
            'Buy Date': buy_date.strftime('%Y-%m-%d') if buy_date else None,
            'Current Price': current_price
        }

        res = update_stock_holding(user_id, row_idx, holding_data)
        if res:
            clear_user_caches(user_id)
            return jsonify({"status": "success", "message": f"Stock '{company_name}' updated successfully in Supabase!"})
        else:
            return jsonify({"status": "error", "message": "Failed to update stock in Supabase."}), 500
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/api/stock/save-buy-date', methods=['POST'])
def save_stock_buy_date():
    user_id = get_current_user_id()
    if not user_id:
        return jsonify({"status": "error", "message": "Unauthorized"}), 401
    data = request.json
    row_idx = (data.get("row_idx") or
               data.get("holding_id") or
               data.get("id"))
    buy_date_raw = data.get("Buy Date")
    try:
        from lib.supabase_data import update_stock_holding
        buy_date = parse_date(buy_date_raw) if buy_date_raw else None
        
        res = update_stock_holding(user_id, row_idx, {
            'Buy Date': buy_date.strftime('%Y-%m-%d') if buy_date else None
        })
        if res:
            clear_user_caches(user_id)
            return jsonify({"status": "success", "message": "Buy date saved successfully to Supabase!"})
        else:
            return jsonify({"status": "error", "message": "Failed to save buy date"}), 500
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/api/stock/save-sector', methods=['POST'])
def save_stock_sector():
    user_id = get_current_user_id()
    if not user_id:
        return jsonify({"status": "error", "message": "Unauthorized"}), 401
    data = request.json
    row_idx = (data.get("row_idx") or
               data.get("holding_id") or
               data.get("id"))
    sector = data.get("Sector", "").strip()
    if not sector:
        return jsonify({"status": "error", "message": "Sector is required."}), 400
    try:
        from lib.supabase_data import update_stock_holding
        res = update_stock_holding(user_id, row_idx, {
            'Sector': sector
        })
        if res:
            clear_user_caches(user_id)
            return jsonify({"status": "success", "message": f"Sector updated successfully to '{sector}' in Supabase!"})
        else:
            return jsonify({"status": "error", "message": "Failed to save sector"}), 500
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/api/stock/delete', methods=['POST'])
def delete_stock():
    user_id = get_current_user_id()
    if not user_id:
        return jsonify({"status": "error", "message": "Unauthorized"}), 401
    data = request.json
    holding_id = (data.get("holding_id") or
                  data.get("row_idx") or
                  data.get("id"))
    if not holding_id:
        return jsonify({"status": "error", "message": "holding_id required"}), 400
    try:
        from lib.supabase_data import delete_stock_holding
        success = delete_stock_holding(user_id, holding_id)
        if success:
            clear_user_caches(user_id)
            return jsonify({"status": "success", "message": "Stock holding deleted from Supabase!"})
        else:
            return jsonify({"status": "error", "message": "Failed to delete stock from Supabase."}), 500
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/api/stock/delete-all', methods=['POST'])
def delete_all_stocks():
    user_id = get_current_user_id()
    if not user_id:
        return jsonify({"status": "error", "message": "Unauthorized"}), 401
    try:
        supabase = get_supabase_admin()
        supabase.table('stock_holdings').delete().eq('user_id', user_id).execute()
        clear_user_caches(user_id)
        return jsonify({"status": "success", "message": "All stock holdings deleted successfully from Supabase!"})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/api/mf/add', methods=['POST'])
def add_mf():
    user_id = get_current_user_id()
    if not user_id:
        return jsonify({"status": "error", "message": "Unauthorized"}), 401
    data = request.json
    try:
        from lib.supabase_data import add_mf_holding
        fund_name = data.get("Fund Name", "").strip()
        amc       = data.get("AMC", "Other").strip()
        category  = data.get("Category", "Other").strip()
        sub_cat   = data.get("Sub-Category", "Other").strip()
        units     = float(data.get("Units Held", 0) or 0)
        buy_nav   = float(data.get("Buy NAV", 0) or 0)
        
        current_nav = data.get("Current NAV")
        if current_nav is not None and str(current_nav).strip() != "":
            current_nav = float(current_nav)
        else:
            current_nav = None

        if current_nav is None:
            code = _resolve_mf_scheme_code(fund_name)
            if code:
                try:
                    url = f"https://api.mfapi.in/mf/{code}"
                    r = requests.get(url, timeout=5)
                    if r.status_code == 200:
                        res_data = r.json().get('data', [])
                        if res_data:
                            current_nav = float(res_data[0]['nav'])
                except Exception:
                    pass
            if current_nav is None:
                current_nav = buy_nav

        buy_date_str = data.get("Buy Date", "")
        parsed_d = parse_date(buy_date_str)
        buy_date = parsed_d.strftime('%Y-%m-%d') if parsed_d else None

        holding_data = {
            'Fund Name': fund_name,
            'AMC': amc,
            'Category': category,
            'Sub-Category': sub_cat,
            'Units Held': units,
            'Buy NAV': buy_nav,
            'Current NAV': current_nav,
            'Buy Date': buy_date
        }

        res = add_mf_holding(user_id, holding_data)
        if res:
            clear_user_caches(user_id)
            return jsonify({"status": "success", "message": "Mutual Fund added successfully to Supabase!"})
        else:
            return jsonify({"status": "error", "message": "Failed to add Mutual Fund"}), 500
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/mf/edit', methods=['POST'])
def edit_mf():
    user_id = get_current_user_id()
    if not user_id:
        return jsonify({"status": "error", "message": "Unauthorized"}), 401
    data = request.json
    row_idx = (data.get("row_idx") or
               data.get("holding_id") or
               data.get("id"))
    try:
        from lib.supabase_data import update_mf_holding
        fund_name = data.get("Fund Name", "").strip()
        amc       = data.get("AMC", "Other").strip()
        category  = data.get("Category", "Other").strip()
        sub_cat   = data.get("Sub-Category", "Other").strip()
        units     = float(data.get("Units Held", 0) or 0)
        buy_nav   = float(data.get("Buy NAV", 0) or 0)

        current_nav = data.get("Current NAV")
        if current_nav is not None and str(current_nav).strip() != "":
            current_nav = float(current_nav)
        else:
            current_nav = None

        if current_nav is None:
            code = _resolve_mf_scheme_code(fund_name)
            if code:
                try:
                    url = f"https://api.mfapi.in/mf/{code}"
                    r = requests.get(url, timeout=5)
                    if r.status_code == 200:
                        res_data = r.json().get('data', [])
                        if res_data:
                            current_nav = float(res_data[0]['nav'])
                except Exception:
                    pass
            if current_nav is None:
                current_nav = buy_nav

        buy_date_str = data.get("Buy Date", "")
        parsed_d = parse_date(buy_date_str)
        buy_date = parsed_d.strftime('%Y-%m-%d') if parsed_d else None

        holding_data = {
            'Fund Name': fund_name,
            'AMC': amc,
            'Category': category,
            'Sub-Category': sub_cat,
            'Units Held': units,
            'Buy NAV': buy_nav,
            'Current NAV': current_nav,
            'Buy Date': buy_date
        }

        res = update_mf_holding(user_id, row_idx, holding_data)
        if res:
            clear_user_caches(user_id)
            return jsonify({"status": "success", "message": "Mutual Fund updated successfully in Supabase!"})
        else:
            return jsonify({"status": "error", "message": "Failed to update Mutual Fund"}), 500
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/mf/delete', methods=['POST'])
def delete_mf():
    user_id = get_current_user_id()
    if not user_id:
        return jsonify({"status": "error", "message": "Unauthorized"}), 401
    data = request.json
    holding_id = (data.get("holding_id") or
                  data.get("row_idx") or
                  data.get("id"))
    if not holding_id:
        return jsonify({"status": "error", "message": "holding_id required"}), 400
    try:
        from lib.supabase_data import delete_mf_holding
        success = delete_mf_holding(user_id, holding_id)
        if success:
            clear_user_caches(user_id)
            return jsonify({"status": "success", "message": "Mutual Fund holding deleted from Supabase!"})
        else:
            return jsonify({"status": "error", "message": "Failed to delete mutual fund"}), 500
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/update-prices', methods=['POST'])
def update_prices():
    user_id = get_current_user_id()
    if not user_id:
        return jsonify({"status": "error", "message": "Unauthorized"}), 401
        
    global price_update_state
    if "users" not in price_update_state:
        price_update_state["users"] = {}
        
    user_state = price_update_state["users"].get(user_id, {})
    if user_state.get("status") in ("running", "background_running"):
        return jsonify({"status": "error", "message": "Update already in progress."}), 400
        
    threading.Thread(target=live_price_updater_thread, args=(user_id,), daemon=True).start()
    return jsonify({"status": "success", "message": "Live price update started in the background."})

@app.route('/api/update-status', methods=['GET'])
def update_status():
    user_id = get_current_user_id()
    if not user_id:
        return jsonify({"status": "error", "message": "Unauthorized"}), 401
        
    global price_update_state
    if "users" not in price_update_state:
        price_update_state["users"] = {}
        
    user_state = price_update_state["users"].setdefault(user_id, {
        "status": "idle",
        "logs": [],
        "start_time": 0,
        "current_index": 0,
        "total_tickers": 0,
        "current_ticker": ""
    })
    
    pct = 0
    if user_state["total_tickers"] > 0:
        pct = int((user_state["current_index"] / user_state["total_tickers"]) * 100)
    
    elapsed = 0
    if user_state["start_time"] and user_state["status"] in ("running", "background_running"):
        elapsed = round(time.time() - user_state["start_time"], 1)
        
    last_updated = price_update_state["last_updated_by_profile"].get(user_id)
    
    return jsonify({
        "status": user_state["status"],
        "progress_pct": pct,
        "current_index": user_state["current_index"],
        "total_tickers": user_state["total_tickers"],
        "current_ticker": user_state["current_ticker"],
        "elapsed_seconds": elapsed,
        "last_updated": last_updated,
        "logs": user_state["logs"][-15:]
    })

@app.route('/api/advisor/chat', methods=['POST'])
def advisor_chat():
    user_id = get_current_user_id()
    if not user_id:
        return jsonify({"response": "Please login to use AI Advisor."}), 401

    if not check_plan_limit(user_id, 'ai_advisor'):
        return jsonify({"response": """
            <div style='text-align:center;padding:2rem;'>
              <i class='fa-solid fa-lock' style='font-size:2rem;color:var(--accent);margin-bottom:1rem;display:block;'></i>
              <strong>AI Advisor is a Pro feature</strong>
              <p style='color:var(--text-secondary);margin-top:0.5rem;'>Upgrade to Pro to get personalised AI-powered investment advice.</p>
              <a href='/account' style='display:inline-block;margin-top:1rem;padding:8px 20px;background:var(--accent);color:white;border-radius:6px;text-decoration:none;'>Upgrade to Pro</a>
            </div>"""})

    try:
        data   = request.json or {}
        prompt = data.get("prompt", "").strip().lower()
        if not prompt:
            return jsonify({"response": "I didn't receive any investment query. Please write what's on your mind!"})

        # ── Pull live data directly from Supabase ─────────────────────────────
        from lib.supabase_data import get_user_stock_holdings, get_user_mf_holdings
        raw_stocks = get_user_stock_holdings(user_id)
        raw_mfs    = get_user_mf_holdings(user_id)

        # ── Build enriched stock list ─────────────────────────────────────────
        stocks = []
        for s in raw_stocks:
            scrip     = str(s.get('scrip_name') or '').strip()
            qty       = float(s.get('quantity') or 0)
            buy_p     = float(s.get('buy_price') or 0)
            cur_p     = float(s.get('current_price') or buy_p or 0)
            invested  = round(qty * buy_p, 2)
            cur_val   = round(qty * cur_p, 2)
            pnl       = round(cur_val - invested, 2)
            ret_pct   = round((pnl / invested * 100) if invested > 0 else 0, 2)
            stocks.append({
                "Scrip Name":     scrip,
                "Sector":         str(s.get('sector') or 'Other').strip(),
                "Invested Value": invested,
                "Current Value":  cur_val,
                "P&L":            pnl,
                "Return %":       ret_pct,
            })

        # ── MF totals ─────────────────────────────────────────────────────────
        mf_invested = sum(float(m.get('units_held') or 0) * float(m.get('buy_nav') or 0)     for m in raw_mfs)
        mf_val      = sum(float(m.get('units_held') or 0) * float(m.get('current_nav') or m.get('buy_nav') or 0) for m in raw_mfs)

        # ── Portfolio-level summary ───────────────────────────────────────────
        total_stock_val  = sum(s["Current Value"] for s in stocks)
        total_val        = total_stock_val + mf_val
        total_invested   = sum(s["Invested Value"] for s in stocks) + mf_invested
        total_pnl        = round(total_val - total_invested, 2)
        total_ret        = round((total_pnl / total_invested * 100) if total_invested > 0 else 0, 2)
        active_stocks    = len(stocks)
        active_mfs       = len(raw_mfs)

        # ── Concentration: top-5 by current value ────────────────────────────
        by_val  = sorted(stocks, key=lambda x: x["Current Value"], reverse=True)
        top5    = by_val[:5]
        top5_wt = round((sum(s["Current Value"] for s in top5) / total_val * 100) if total_val > 0 else 0, 2)
        top1    = top5[0] if top5 else None
        top2    = top5[1] if len(top5) > 1 else None
        top1_wt = round((top1["Current Value"] / total_val * 100) if top1 and total_val > 0 else 0, 2)
        top2_wt = round((top2["Current Value"] / total_val * 100) if top2 and total_val > 0 else 0, 2)

        # ── Best / worst performers ───────────────────────────────────────────
        worst3 = sorted(stocks, key=lambda x: x["Return %"])[:3]
        best3  = sorted(stocks, key=lambda x: x["Return %"], reverse=True)[:3]
        best_str  = ", ".join(f"<strong>{s['Scrip Name']} (+{s['Return %']:.1f}%)</strong>" for s in best3)  if best3  else "N/A"
        worst_str = ", ".join(f"<strong>{s['Scrip Name']} ({s['Return %']:.1f}%)</strong>"  for s in worst3) if worst3 else "N/A"

        # ── Sector gaps ───────────────────────────────────────────────────────
        covered = {s["Sector"].lower() for s in stocks}
        missing_sectors = [k for k in ["Technology", "Healthcare", "Consumer Staples", "Financials"]
                           if not any(k.lower() in c for c in covered)]

        # ── Portfolio beta proxy ──────────────────────────────────────────────
        _BETA = {"Technology": 1.1, "Financials": 1.0, "Banking & Finance": 1.0,
                 "Energy": 0.9, "Materials": 1.2, "Auto": 1.15,
                 "Consumer Discretionary": 1.05, "Consumer Staples": 0.7,
                 "Healthcare": 0.75, "Industrials": 1.1, "Communication": 1.0, "ETF": 1.0}
        wb = sum(_BETA.get(s["Sector"], 1.0) * s["Current Value"] for s in stocks)
        portfolio_beta = round((wb / total_stock_val) if total_stock_val > 0 else 1.0, 2)

        # ═══════════════════════════════════════════════════════════════════════
        # RESPONSE BRANCHES
        # ═══════════════════════════════════════════════════════════════════════
        if "concentration" in prompt or "risk" in prompt or "audit" in prompt:
            top5_names = " + ".join(s["Scrip Name"] for s in top5) or "N/A"
            rows = ""
            if top1:
                flag = " — exceeds the safe 10% institutional ceiling" if top1_wt > 10 else ""
                rows += f"<li><strong>Top Holding ({top1['Scrip Name']}):</strong> <strong>{top1_wt:.2f}%</strong> of AUM (₹{top1['Current Value']:,.2f}){flag}.</li>\n"
            if top2:
                rows += f"<li><strong>2nd Holding ({top2['Scrip Name']}):</strong> <strong>{top2_wt:.2f}%</strong> of AUM (₹{top2['Current Value']:,.2f}).</li>\n"
            rows += f"<li><strong>Top-5 Combined ({top5_names}):</strong> <strong>{top5_wt:.2f}%</strong> of total portfolio.</li>\n"
            gap_note = (f"<strong>Sector Gaps:</strong> Little-to-no exposure in <em>{', '.join(missing_sectors)}</em>. Consider diversifying.<br><br>"
                        if missing_sectors else "")
            response = f"""
<span style="font-weight:700;color:var(--danger);font-size:1rem;display:block;margin-bottom:0.5rem;"><i class="fa-solid fa-triangle-exclamation"></i> Portfolio Concentration Risk Audit</span>
Our diagnostic scan reviewed your <strong>{active_stocks} holdings</strong> (₹{total_val:,.2f} AUM):
<ul>{rows}</ul>
{gap_note}<span style="color:var(--accent);font-weight:600;">Advisory Opinion:</span>
If any single position exceeds 15% of your portfolio, consider trimming and redistributing into under-represented sectors.
A portfolio beta of ~1.0 (yours: {portfolio_beta}) with diversified sector coverage reduces drawdown risk significantly.
"""

        elif "stress" in prompt or "scenario" in prompt or "crisis" in prompt or "correction" in prompt:
            response = f"""
<span style="font-weight:700;color:var(--warning);font-size:1rem;display:block;margin-bottom:0.5rem;"><i class="fa-solid fa-bolt"></i> Macro Scenario Stress Test (12-Month Outlook)</span>
Stress-testing your <strong>{active_stocks}-stock + {active_mfs}-MF</strong> portfolio (₹{total_val:,.2f}, beta ≈ {portfolio_beta}):
<ol>
  <li><strong>Bull Run (+20% Nifty 50):</strong><br>Expected: <span style="color:var(--success);font-weight:600;">+{20*portfolio_beta:.1f}% → ₹{total_val*(1+0.20*portfolio_beta):,.2f}</span>. High-beta holdings amplify gains.</li>
  <li><strong>Normal Correction (-10%):</strong><br>Expected: <span style="color:var(--danger);font-weight:600;">{-10*portfolio_beta:.1f}% → -₹{total_val*0.10*portfolio_beta:,.2f}</span>. Defensive gaps will hurt relative performance.</li>
  <li><strong>Severe Bear Market (-25%):</strong><br>Expected: <span style="color:var(--danger);font-weight:600;">{-25*portfolio_beta:.1f}% → -₹{total_val*0.25*portfolio_beta:,.2f}</span>. Cyclical names will drag heavily.</li>
  <li><strong>Systemic Crisis (-40%):</strong><br>Expected: <span style="color:var(--danger);font-weight:700;">{-40*portfolio_beta:.1f}% → -₹{total_val*0.40*portfolio_beta:,.2f}</span>. Speculative small-caps face severe erosion.</li>
</ol>
<span style="color:var(--text-secondary);font-size:0.85rem;">*Returns scaled by estimated portfolio beta ({portfolio_beta}). Actual outcomes depend on individual stock fundamentals.*</span>
"""

        elif "rebalance" in prompt or "plan" in prompt or "recommendation" in prompt:
            overweight  = [s for s in stocks if total_val > 0 and (s["Current Value"] / total_val * 100) > 15]
            deep_losers = [s for s in stocks if s["Return %"] < -20]
            ow_lines = "".join(
                f"&bull; Trim <strong>{s['Scrip Name']}</strong> ({s['Current Value']/total_val*100:.1f}% → ~10%)"
                f" — frees ≈ ₹{(s['Current Value'] - total_val*0.10):,.0f}.<br>"
                for s in overweight
            ) or "&bull; No single position exceeds 15% — allocation looks balanced.<br>"
            loser_lines = "".join(
                f"&bull; Consider exiting <strong>{s['Scrip Name']}</strong> ({s['Return %']:.1f}%) — tax-loss harvest &amp; recycle capital.<br>"
                for s in deep_losers
            ) or "&bull; No deep losers (&gt;-20%) detected — good job managing downside.<br>"
            response = f"""
<span style="font-weight:700;color:var(--accent);font-size:1rem;display:block;margin-bottom:0.5rem;"><i class="fa-solid fa-arrows-spin"></i> Actionable Rebalancing &amp; Capital Recycling Plan</span>
Based on your <strong>{active_stocks} stocks + {active_mfs} MFs</strong> (₹{total_val:,.2f} total):
<br><br>
<strong>Step 1 — Trim Overweight Positions:</strong><br>{ow_lines}
<br><strong>Step 2 — Exit Deep Underperformers:</strong><br>{loser_lines}
<br><strong>Step 3 — Deploy Freed Capital:</strong><br>
&bull; <strong>Quality Financials:</strong> HDFC Bank, ICICI Bank for stable compounding.<br>
&bull; <strong>IT / Technology:</strong> TCS or Infosys for a structural growth moat.<br>
&bull; <strong>Defensives / FMCG:</strong> ITC, Sun Pharma, or HUL to reduce portfolio beta.<br>
&bull; <strong>Index ETFs:</strong> NiftyBees / MON100 to anchor broad market exposure.
"""

        elif "worst" in prompt or "perform" in prompt or "loss" in prompt or "poor" in prompt:
            rows = ""
            for ws in worst3:
                label = f"Loss ₹{abs(ws['P&L']):,.2f}" if ws["P&L"] < 0 else f"Gain ₹{ws['P&L']:,.2f}"
                rows += f"<li><strong>{ws['Scrip Name']}:</strong> <span style='color:var(--danger);font-weight:600;'>{ws['Return %']:.1f}%</span> ({label} on ₹{ws['Invested Value']:,.2f})</li>"
            response = f"""
<span style="font-weight:700;color:var(--danger);font-size:1rem;display:block;margin-bottom:0.5rem;"><i class="fa-solid fa-circle-down"></i> Underperforming Assets Diagnostic</span>
Your 3 worst-performing holdings:
<ol>{rows}</ol>
<strong>Advisory Verdict:</strong><br>
For holdings down &gt;30%, evaluate if the fundamental thesis has broken. If yes, tax-loss harvest and recycle into quality compounders.
For 10–30% drawdowns, hold if conviction remains — average-down only for high-quality names with strong balance sheets.
"""

        elif "best" in prompt or "top" in prompt or "winner" in prompt or "gain" in prompt:
            rows = ""
            for bs in best3:
                rows += f"<li><strong>{bs['Scrip Name']}:</strong> <span style='color:var(--success);font-weight:600;'>+{bs['Return %']:.1f}%</span> (Gain ₹{bs['P&L']:,.2f} on ₹{bs['Invested Value']:,.2f})</li>"
            response = f"""
<span style="font-weight:700;color:var(--success);font-size:1rem;display:block;margin-bottom:0.5rem;"><i class="fa-solid fa-circle-up"></i> Top Performing Holdings</span>
Your 3 best-performing holdings:
<ol>{rows}</ol>
<strong>Advisory Verdict:</strong><br>
For holdings up &gt;50%, consider booking partial profits (25–30% of position) to lock in gains and reduce concentration risk.
Let the remaining position run with a trailing stop-loss strategy.
"""

        else:
            response = f"""
<span style="font-weight:700;color:var(--accent);font-size:1rem;display:block;margin-bottom:0.5rem;"><i class="fa-solid fa-user-doctor"></i> CFA Custom Wealth Consultation</span>
I have analysed your query: "<em>{data.get('prompt')}</em>"
<br><br>
<strong>Portfolio Snapshot:</strong> {active_stocks} stocks + {active_mfs} MFs &nbsp;|&nbsp; ₹{total_val:,.2f} total value &nbsp;|&nbsp; P&amp;L: ₹{total_pnl:,.2f} ({'+' if total_ret >= 0 else ''}{total_ret:.2f}%)
<br><br>
Top performers: {best_str}<br>
Underperformers: {worst_str}
<br><br>
<strong>Advisor's Recommendation:</strong><br>
&bull; Type <strong>"Concentration Audit"</strong> — review position sizing &amp; sector risk.<br>
&bull; Type <strong>"Stress Test"</strong> — simulate portfolio under market crashes.<br>
&bull; Type <strong>"Rebalancing Plan"</strong> — get concrete buy/sell/trim steps.<br>
&bull; Type <strong>"Worst Performers"</strong> — triage your biggest losers.<br>
&bull; Type <strong>"Best Performers"</strong> — review top gainers &amp; profit-booking strategy.
"""

        return jsonify({"response": response})
    except Exception as e:
        return jsonify({"response": f"Wealth Advisor API error: {str(e)}"}), 500

@app.route('/api/profiles', methods=['GET'])
def get_profiles():
    try:
        meta_file = "profiles_config.json"
        load_profile_globals()  # ensures init
        with open(meta_file, "r") as f:
            meta = json.load(f)
        profiles_list = []
        for p_name, p_info in meta.get("profiles", {}).items():
            profiles_list.append({
                "name": p_name,
                "has_pin": p_info.get("pin_hash") is not None
            })
        return jsonify({
            "active_profile": meta.get("active_profile", "Default Portfolio"),
            "profiles": profiles_list
        })
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/profiles/select', methods=['POST'])
def select_profile():
    try:
        data = request.json or {}
        name = data.get("profile_name", "").strip()
        pin = data.get("pin", "")
        if not name:
            return jsonify({"status": "error", "message": "Profile name is required."}), 400
        
        meta_file = "profiles_config.json"
        if not os.path.exists(meta_file):
            return jsonify({"status": "error", "message": "Profiles meta config not initialized."}), 400
            
        with open(meta_file, "r") as f:
            meta = json.load(f)
            
        profile_data = meta.get("profiles", {}).get(name)
        if not profile_data:
            return jsonify({"status": "error", "message": f"Profile '{name}' does not exist."}), 404
            
        stored_hash = profile_data.get("pin_hash")
        if stored_hash:
            if not pin or hash_pin(pin) != stored_hash:
                return jsonify({"status": "error", "message": "Incorrect PIN. Access denied."}), 401
            
        meta["active_profile"] = name
        with open(meta_file, "w") as f:
            json.dump(meta, f, indent=4)
            
        load_profile_globals()
        return jsonify({"status": "success", "message": f"Switched to profile '{name}' successfully!", "active_profile": name})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/profiles/create', methods=['POST'])
def create_profile():
    try:
        data = request.json or {}
        name = data.get("profile_name", "").strip()
        if not name:
            return jsonify({"status": "error", "message": "Profile name is required."}), 400
            
        # Sanitize safe directory name
        safe_name = "".join([c if c.isalnum() else "_" for c in name]).lower()
        profile_dir = os.path.join("profiles", safe_name)
        os.makedirs(profile_dir, exist_ok=True)
        
        new_excel_path = os.path.join(profile_dir, "portfolio.xlsx").replace('\\', '/')
        new_config_path = os.path.join(profile_dir, "config.json").replace('\\', '/')
        
        master_excel = "Stocks & MF Analysis_V3.xlsx"
        master_config = "config.json"
        
        if not os.path.exists(new_excel_path):
            if os.path.exists(master_excel):
                shutil.copy(master_excel, new_excel_path)
                # Wipe pre-populated stock and mutual fund holdings (rows 4 onwards)
                try:
                    wb = openpyxl.load_workbook(new_excel_path)
                    if "📥 Stock Data Input" in wb.sheetnames:
                        ws_s = wb["📥 Stock Data Input"]
                        ws_s.delete_rows(4, 1000)
                    if "📥 MF Data Input" in wb.sheetnames:
                        ws_m = wb["📥 MF Data Input"]
                        ws_m.delete_rows(4, 1000)
                    wb.save(new_excel_path)
                except Exception as e:
                    print("Error preparing clean template spreadsheet:", e)
            else:
                return jsonify({"status": "error", "message": "Master template spreadsheet not found."}), 500
                
        if not os.path.exists(new_config_path):
            if os.path.exists(master_config):
                shutil.copy(master_config, new_config_path)
            else:
                default_config = {
                    "theme": "emerald",
                    "display_columns": {
                        "stocks": ["Scrip Name", "Exchange", "Sector", "Qty", "Buy Price", "Buy Date", "Current Price", "Invested Value", "Current Value", "P&L", "Return %", "Tax Flag"],
                        "mf": ["Fund Name", "Category", "Units Held", "Invested Value", "Current Value", "P&L", "Return %", "XIRR %", "Holding Period", "Tax Flag"]
                    },
                    "widgets": {
                        "sector_allocation": True,
                        "asset_allocation": True,
                        "top_performers": True,
                        "portfolio_signals": True,
                        "dividend_yield": True,
                        "risk_alerts": True
                    },
                    "custom_alerts": {
                        "high_concentration_pct": 20,
                        "high_loss_pct": 10
                    }
                }
                with open(new_config_path, "w") as f:
                    json.dump(default_config, f, indent=4)

                    
        meta_file = "profiles_config.json"
        if not os.path.exists(meta_file):
            load_profile_globals()
            
        with open(meta_file, "r") as f:
            meta = json.load(f)
            
        pin = data.get("pin", "").strip()
        profile_info = {
            "excel_path": new_excel_path,
            "config_path": new_config_path
        }
        if pin:
            profile_info["pin_hash"] = hash_pin(pin)
            
        meta["profiles"][name] = profile_info
        meta["active_profile"] = name
        with open(meta_file, "w") as f:
            json.dump(meta, f, indent=4)
            
        load_profile_globals()
        return jsonify({
            "status": "success", 
            "message": f"Profile '{name}' created and loaded successfully!",
            "active_profile": name
        })
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/profiles/delete', methods=['POST'])
def delete_profile():
    try:
        data = request.json or {}
        name = data.get("profile_name", "").strip()
        if not name:
            return jsonify({"status": "error", "message": "Profile name is required."}), 400
            
        if name == "Default Portfolio":
            return jsonify({"status": "error", "message": "The Default Portfolio profile cannot be deleted."}), 400
            
        meta_file = "profiles_config.json"
        if not os.path.exists(meta_file):
            return jsonify({"status": "error", "message": "Profiles meta config not initialized."}), 400
            
        with open(meta_file, "r") as f:
            meta = json.load(f)
            
        if name not in meta.get("profiles", {}):
            return jsonify({"status": "error", "message": f"Profile '{name}' does not exist."}), 404
            
        # Verify PIN if profile is locked
        pin = data.get("pin", "")
        stored_hash = meta["profiles"][name].get("pin_hash")
        if stored_hash:
            if not pin or hash_pin(pin) != stored_hash:
                return jsonify({"status": "error", "message": "Incorrect PIN. Access denied."}), 401
            
        # Safely delete profile folder under 'profiles/'
        safe_name = "".join([c if c.isalnum() else "_" for c in name]).lower()
        profile_dir = os.path.join("profiles", safe_name)
        if os.path.exists(profile_dir):
            # Double check to prevent deleting anything outside of the 'profiles' subfolder
            parent_dir = os.path.basename(os.path.dirname(os.path.abspath(profile_dir)))
            if parent_dir == "profiles" or "profiles" in os.path.split(profile_dir)[0]:
                try:
                    shutil.rmtree(profile_dir)
                except Exception as e:
                    print(f"Error deleting profile directory {profile_dir}: {e}")
                    
        # Remove from configuration dictionary
        del meta["profiles"][name]
        
        # If the deleted profile was currently active, switch back to Default Portfolio
        active_switched = False
        if meta.get("active_profile") == name:
            meta["active_profile"] = "Default Portfolio"
            active_switched = True
            
        with open(meta_file, "w") as f:
            json.dump(meta, f, indent=4)
            
        load_profile_globals()
        return jsonify({
            "status": "success",
            "message": f"Profile '{name}' deleted successfully!",
            "active_profile": meta["active_profile"],
            "active_switched": active_switched
        })
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/profiles/rename', methods=['POST'])
def rename_profile():
    try:
        data = request.json or {}
        old_name = data.get("old_name", "").strip()
        new_name = data.get("new_name", "").strip()

        if not old_name or not new_name:
            return jsonify({"status": "error", "message": "Both old and new profile names are required."}), 400

        if old_name == "Default Portfolio":
            return jsonify({"status": "error", "message": "The Default Portfolio profile cannot be renamed."}), 400

        if old_name == new_name:
            return jsonify({"status": "error", "message": "New name is the same as the current name."}), 400

        meta_file = "profiles_config.json"
        if not os.path.exists(meta_file):
            return jsonify({"status": "error", "message": "Profiles meta config not initialized."}), 400

        with open(meta_file, "r") as f:
            meta = json.load(f)

        if old_name not in meta.get("profiles", {}):
            return jsonify({"status": "error", "message": f"Profile '{old_name}' does not exist."}), 404

        # Verify PIN if profile is locked
        pin = data.get("pin", "")
        stored_hash = meta["profiles"][old_name].get("pin_hash")
        if stored_hash:
            if not pin or hash_pin(pin) != stored_hash:
                return jsonify({"status": "error", "message": "Incorrect PIN. Access denied."}), 401

        if new_name in meta.get("profiles", {}):
            return jsonify({"status": "error", "message": f"A profile named '{new_name}' already exists."}), 400

        # Rename the profile directory on disk if it exists inside profiles/
        old_safe = "".join([c if c.isalnum() else "_" for c in old_name]).lower()
        new_safe = "".join([c if c.isalnum() else "_" for c in new_name]).lower()
        old_dir = os.path.join("profiles", old_safe)
        new_dir = os.path.join("profiles", new_safe)

        if os.path.exists(old_dir) and not os.path.exists(new_dir):
            os.rename(old_dir, new_dir)

        # Update the excel/config paths in the registry to use new directory name
        profile_data = meta["profiles"][old_name]
        if old_safe in profile_data.get("excel_path", ""):
            profile_data["excel_path"] = profile_data["excel_path"].replace(old_safe, new_safe)
        if old_safe in profile_data.get("config_path", ""):
            profile_data["config_path"] = profile_data["config_path"].replace(old_safe, new_safe)

        # Re-insert under new key, remove old key
        meta["profiles"][new_name] = profile_data
        del meta["profiles"][old_name]

        # Update active_profile pointer if the renamed profile was active
        if meta.get("active_profile") == old_name:
            meta["active_profile"] = new_name

        with open(meta_file, "w") as f:
            json.dump(meta, f, indent=4)

        load_profile_globals()
        return jsonify({
            "status": "success",
            "message": f"Profile renamed from '{old_name}' to '{new_name}' successfully!",
            "active_profile": meta["active_profile"]
        })
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/profiles/set-pin', methods=['POST'])
def set_profile_pin():
    try:
        data = request.json or {}
        current_pin = data.get("current_pin", "")
        new_pin = data.get("new_pin", "").strip()
        
        meta_file = "profiles_config.json"
        if not os.path.exists(meta_file):
            return jsonify({"status": "error", "message": "Profiles meta config not initialized."}), 400
            
        with open(meta_file, "r") as f:
            meta = json.load(f)
            
        active = meta.get("active_profile", "Default Portfolio")
        profile_data = meta.get("profiles", {}).get(active)
        if not profile_data:
            return jsonify({"status": "error", "message": "Active profile data not found."}), 404
            
        stored_hash = profile_data.get("pin_hash")
        
        # Verify current PIN if one is set
        if stored_hash:
            if not current_pin or hash_pin(current_pin) != stored_hash:
                return jsonify({"status": "error", "message": "Incorrect current PIN."}), 401
                
        # If new_pin is empty, remove PIN protection
        if not new_pin:
            if "pin_hash" in profile_data:
                del profile_data["pin_hash"]
            msg = "PIN protection disabled successfully."
        else:
            profile_data["pin_hash"] = hash_pin(new_pin)
            msg = "PIN updated successfully."
            
        with open(meta_file, "w") as f:
            json.dump(meta, f, indent=4)
            
        return jsonify({"status": "success", "message": msg})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/profiles/set-pin-by-name', methods=['POST'])
def set_profile_pin_by_name():
    try:
        data = request.json or {}
        name = data.get("profile_name", "").strip()
        current_pin = data.get("current_pin", "")
        new_pin = data.get("new_pin", "").strip()
        
        if not name:
            return jsonify({"status": "error", "message": "Profile name is required."}), 400
            
        meta_file = "profiles_config.json"
        if not os.path.exists(meta_file):
            return jsonify({"status": "error", "message": "Profiles config not initialized."}), 400
            
        with open(meta_file, "r") as f:
            meta = json.load(f)
            
        profile_data = meta.get("profiles", {}).get(name)
        if not profile_data:
            return jsonify({"status": "error", "message": f"Profile '{name}' does not exist."}), 404
            
        stored_hash = profile_data.get("pin_hash")
        
        # Verify current PIN if one is set
        if stored_hash:
            if not current_pin or hash_pin(current_pin) != stored_hash:
                return jsonify({"status": "error", "message": "Incorrect current PIN. Access denied."}), 401
                
        # If new_pin is empty, remove PIN protection
        if not new_pin:
            if "pin_hash" in profile_data:
                del profile_data["pin_hash"]
            msg = f"PIN protection disabled for profile '{name}'."
        else:
            profile_data["pin_hash"] = hash_pin(new_pin)
            msg = f"PIN updated successfully for profile '{name}'."
            
        with open(meta_file, "w") as f:
            json.dump(meta, f, indent=4)
            
        return jsonify({"status": "success", "message": msg})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

if __name__ == "__main__":
    if os.environ.get("WERKZEUG_RUN_MAIN") == "true" or not app.debug:
        threading.Thread(target=automatic_price_updater_loop, daemon=True).start()
    app.run(debug=True, port=5000)
